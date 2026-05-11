from __future__ import annotations

import argparse
import json
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import cast
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException

from logic_controller import LogicController

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_ARTIFACT_DIR = BASE_DIR.parent / "artifacts" / "perf"


class WorkbookParseError(RuntimeError):
    pass


def _now_utc() -> str:
    return datetime.now(tz=timezone.utc).isoformat()


def _perf_ms(start_ns: int, end_ns: int) -> float:
    return round((end_ns - start_ns) / 1_000_000.0, 3)


def _format_adapter_output(result: dict[str, object]) -> tuple[str, str]:
    status = result.get("status")
    if status == "TM_HIT":
        tm_match = cast(dict[str, object], result.get("tm_match", {}))
        target = tm_match.get("target", "")
        if not isinstance(target, str):
            raise RuntimeError("invalid TM_HIT payload: tm_match.target is not a string")
        return "TM_HIT", f"[TM] {target}"

    if status == "TERMS_FOUND":
        terms = cast(list[dict[str, object]], result.get("terms", []))
        ordered_terms = sorted(
            terms,
            key=lambda entry: int(cast(int, entry.get("start_index", 0))),
        )
        rendered_terms: list[str] = []
        for term in ordered_terms:
            source_term = term.get("source_term", "")
            target_term = term.get("target_term", "")
            if not isinstance(source_term, str) or not isinstance(target_term, str):
                raise RuntimeError("invalid TERMS_FOUND payload: source_term/target_term must be strings")
            rendered_terms.append(f"{source_term}->{target_term}")
        return "TERMS_FOUND", f"[Terms] {', '.join(rendered_terms)}"

    if status == "NO_MATCH":
        return "NO_MATCH", "[No Match]"

    raise RuntimeError(f"unexpected status from LogicController: {status!r}")


def _build_input_workbook_from_workload(
    workload_path: Path,
    input_xlsx: Path,
    source_column: str,
    max_rows: int | None,
) -> int:
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise RuntimeError("failed to initialize workbook active sheet")
    sheet: Worksheet = active_sheet
    row_count = 0

    with workload_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            payload = line.strip()
            if not payload:
                continue
            row = cast(dict[str, object], json.loads(payload))
            source = row.get("source")
            if not isinstance(source, str) or not source.strip():
                continue
            row_count += 1
            sheet[f"{source_column}{row_count}"] = source.strip()
            if max_rows is not None and row_count >= max_rows:
                break

    input_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(input_xlsx)
    return row_count


def run_file_mode_benchmark(
    input_xlsx: Path,
    output_xlsx: Path,
    timing_json: Path,
    source_column: str,
    target_column: str,
    sheet_name: str | None,
    max_rows: int | None,
) -> Path:
    segment_starts = time.perf_counter_ns()
    try:
        workbook = load_workbook(filename=input_xlsx)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        raise WorkbookParseError(f"failed to parse workbook '{input_xlsx}': {exc}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"worksheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        active_sheet = workbook.active
        if active_sheet is None:
            raise RuntimeError("input workbook has no active sheet")
        sheet = active_sheet

    source_column_index = column_index_from_string(source_column)

    source_rows: list[tuple[int, str]] = []
    for row_index in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=source_column_index).value
        if isinstance(cell_value, str):
            clean_text = cell_value.strip()
            if clean_text:
                source_rows.append((row_index, clean_text))
        elif cell_value is not None:
            source_rows.append((row_index, str(cell_value).strip()))
        if max_rows is not None and len(source_rows) >= max_rows:
            break

    load_done = time.perf_counter_ns()
    if not source_rows:
        raise RuntimeError("no source rows found in workbook source column")

    init_start = time.perf_counter_ns()
    controller = LogicController()
    init_done = time.perf_counter_ns()

    compute_start = time.perf_counter_ns()
    formatted_outputs: list[tuple[int, str, str]] = []
    status_counts: Counter[str] = Counter()
    for row_index, source_text in source_rows:
        result = controller.get_suggestions(source_text)
        status, output_text = _format_adapter_output(result)
        status_counts[status] += 1
        formatted_outputs.append((row_index, status, output_text))
    compute_done = time.perf_counter_ns()

    write_start = time.perf_counter_ns()
    for row_index, _, output_text in formatted_outputs:
        sheet[f"{target_column}{row_index}"] = output_text
    write_done = time.perf_counter_ns()

    save_start = time.perf_counter_ns()
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    save_done = time.perf_counter_ns()

    segments = {
        "load_xlsx": _perf_ms(segment_starts, load_done),
        "init_engines": _perf_ms(init_start, init_done),
        "compute_rows": _perf_ms(compute_start, compute_done),
        "write_cells": _perf_ms(write_start, write_done),
        "save_xlsx": _perf_ms(save_start, save_done),
    }
    payload = {
        "harness": "openpyxl_file_mode",
        "generated_at": _now_utc(),
        "input_xlsx": str(input_xlsx.resolve()),
        "output_xlsx": str(output_xlsx.resolve()),
        "sheet": sheet.title,
        "rows_processed": len(formatted_outputs),
        "source_column": source_column,
        "target_column": target_column,
        "timings_ms": segments,
        "status_counts": dict(status_counts),
    }
    timing_json.parent.mkdir(parents=True, exist_ok=True)
    _ = timing_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return timing_json


def main() -> int:
    parser = argparse.ArgumentParser(description="openpyxl file-mode benchmark harness")
    _ = parser.add_argument("--input-xlsx", type=Path, required=True, help="Input workbook to process")
    _ = parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help="Result workbook path (default: artifacts/perf/openpyxl_result_<ts>.xlsx)",
    )
    _ = parser.add_argument(
        "--timing-json",
        type=Path,
        default=None,
        help="Timing artifact path (default: artifacts/perf/openpyxl_filemode_<ts>.json)",
    )
    _ = parser.add_argument(
        "--artifact-dir",
        type=Path,
        default=DEFAULT_ARTIFACT_DIR,
        help="Artifact directory for default output paths",
    )
    _ = parser.add_argument("--sheet", type=str, default=None, help="Worksheet name (default: active)")
    _ = parser.add_argument("--source-column", type=str, default="A", help="Input source column (default: A)")
    _ = parser.add_argument("--target-column", type=str, default="B", help="Output target column (default: B)")
    _ = parser.add_argument("--max-rows", type=int, default=None, help="Optional max rows to process")
    _ = parser.add_argument(
        "--build-input-from-workload",
        type=Path,
        default=None,
        help="Optional workload jsonl path; if set, creates --input-xlsx before benchmarking",
    )
    args = parser.parse_args()

    artifact_dir = cast(Path, args.artifact_dir).expanduser().resolve()
    artifact_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(tz=timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    input_xlsx = cast(Path, args.input_xlsx).expanduser().resolve()
    output_xlsx = cast(Path | None, args.output_xlsx)
    timing_json = cast(Path | None, args.timing_json)
    final_output = output_xlsx.expanduser().resolve() if output_xlsx else artifact_dir / f"openpyxl_result_{stamp}.xlsx"
    final_timing = timing_json.expanduser().resolve() if timing_json else artifact_dir / f"openpyxl_filemode_{stamp}.json"

    source_column = cast(str, args.source_column).strip().upper()
    target_column = cast(str, args.target_column).strip().upper()
    if not source_column or not target_column:
        print("benchmark-error: source/target columns must be non-empty")
        return 1

    build_workload = cast(Path | None, args.build_input_from_workload)
    max_rows = cast(int | None, args.max_rows)

    try:
        if build_workload is not None:
            workload_path = build_workload.expanduser().resolve()
            created_rows = _build_input_workbook_from_workload(
                workload_path=workload_path,
                input_xlsx=input_xlsx,
                source_column=source_column,
                max_rows=max_rows,
            )
            if created_rows == 0:
                print(f"benchmark-error: workload produced zero rows: {workload_path}")
                return 1

        artifact_path = run_file_mode_benchmark(
            input_xlsx=input_xlsx,
            output_xlsx=final_output,
            timing_json=final_timing,
            source_column=source_column,
            target_column=target_column,
            sheet_name=cast(str | None, args.sheet),
            max_rows=max_rows,
        )
    except WorkbookParseError as exc:
        print(f"workbook-parse-error: {exc}")
        return 2
    except Exception as exc:  # noqa: BLE001
        print(f"benchmark-error: {exc}")
        return 1

    print(f"openpyxl benchmark artifact written: {artifact_path}")
    print(f"openpyxl result workbook written: {final_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
