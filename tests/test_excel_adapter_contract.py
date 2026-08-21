from __future__ import annotations

import ast
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from typing import cast
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook

from excel_adapter_openpyxl import run_file_mode_benchmark
import logic_controller
from glossary_engine import GlossaryEngine, GlossaryTerm
from tests.test_tm_activation_journal import SOURCE_BYTES, _first_prepared


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class ExcelAdapterContractTest(unittest.TestCase):
    def test_glossary_engine_batch_apply_rolls_back_second_add_failure(self) -> None:
        engine = GlossaryEngine()
        engine.add_term("existing", "kept", "existing.csv")
        original_root = engine.root
        original_count = engine._term_count
        original_add = GlossaryEngine.add_term
        calls = 0

        def fail_second(staged, source, target, glossary_source, priority=1):
            nonlocal calls
            calls += 1
            if calls == 2:
                raise RuntimeError("injected second add failure")
            return original_add(staged, source, target, glossary_source, priority)

        rows = (
            GlossaryTerm("first", "one", "batch.csv"),
            GlossaryTerm("second", "two", "batch.csv"),
        )
        with patch.object(GlossaryEngine, "add_term", new=fail_second):
            with self.assertRaisesRegex(RuntimeError, "second add failure"):
                engine.apply_terms_atomic(rows)

        self.assertIs(engine.root, original_root)
        self.assertEqual(engine._term_count, original_count)
        self.assertEqual(
            [(hit.source_term, hit.target_term) for hit in engine.extract_terms("existing")],
            [("existing", "kept")],
        )
        self.assertEqual(engine.extract_terms("first second"), [])

    def test_glossary_application_calls_one_atomic_batch_after_terminal(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "terms.csv"
            source.write_text("Source,Target\nfirst,one\nsecond,two\n", encoding="utf-8")
            engine = GlossaryEngine()

            with patch.object(
                engine,
                "apply_terms_atomic",
                wraps=engine.apply_terms_atomic,
            ) as apply_batch:
                logic_controller.load_glossary_file(engine, source)

        apply_batch.assert_called_once()
        terms = apply_batch.call_args.args[0]
        self.assertEqual(
            [(item.source, item.target, item.glossary_source) for item in terms],
            [("first", "one", "terms.csv"), ("second", "two", "terms.csv")],
        )

    def test_glossary_consumer_exception_propagates_without_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "terms.csv"
            source.write_text("first,one\nsecond,two\n", encoding="utf-8")
            engine = GlossaryEngine()
            engine.add_term("existing", "kept", "existing.csv")
            original_root = engine.root
            original_count = engine._term_count

            with patch.object(
                engine,
                "apply_terms_atomic",
                side_effect=RuntimeError("consumer failure"),
            ):
                with self.assertRaisesRegex(RuntimeError, "consumer failure"):
                    logic_controller.load_glossary_file(engine, source)

        self.assertIs(engine.root, original_root)
        self.assertEqual(engine._term_count, original_count)
        self.assertEqual(engine.extract_terms("first second"), [])

    def test_application_does_not_relabel_unexpected_parser_fault(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            csv_source = Path(temporary) / "terms.csv"
            po_source = Path(temporary) / "source.po"
            csv_source.write_text("first,one\n", encoding="utf-8")
            po_source.write_text('msgid "first"\nmsgstr "one"\n', encoding="utf-8")
            surface = Mock()
            surface.open_input.side_effect = AssertionError("programmer fault")

            with patch.object(
                logic_controller,
                "create_parser_application_surface",
                return_value=surface,
            ):
                with self.assertRaisesRegex(AssertionError, "programmer fault"):
                    logic_controller.load_glossary_file(GlossaryEngine(), csv_source)
                with self.assertRaisesRegex(AssertionError, "programmer fault"):
                    logic_controller.load_gettext_source_units(po_source)

    def test_glossary_application_adapter_stages_before_mutating_engine(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "fatal-tail.csv"
            source.write_bytes(b"Source,Target\nkept,value\ninvalid,\xff\n")
            engine = GlossaryEngine()

            with self.assertRaises(logic_controller.GlossaryLoadError) as raised:
                logic_controller.load_glossary_file(engine, source)

        self.assertEqual(raised.exception.code, "PARSER.SOURCE.ENCODING_FAILED")
        self.assertEqual(engine.extract_terms("kept"), [])

    def test_glossary_application_adapter_rejects_retired_xls_explicitly(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "legacy.xls"
            source.write_bytes(b"not-an-xlsx")
            engine = GlossaryEngine()

            with self.assertRaises(logic_controller.GlossaryLoadError) as raised:
                logic_controller.load_glossary_file(engine, source)

        self.assertEqual(raised.exception.code, "PARSER.SELECTION.UNSUPPORTED")
        self.assertEqual(engine.extract_terms("anything"), [])

    def test_all_three_glossary_consumers_share_the_application_adapter(self) -> None:
        controller_source = (PROJECT_ROOT / "logic_controller.py").read_text(
            encoding="utf-8"
        )
        engine_source = (PROJECT_ROOT / "glossary_engine.py").read_text(
            encoding="utf-8"
        )
        translation_source = (PROJECT_ROOT / "translation_runner.py").read_text(
            encoding="utf-8"
        )
        stress_source = (PROJECT_ROOT / "stress_runner.py").read_text(
            encoding="utf-8"
        )

        self.assertIn("load_glossary_file(self.glossary_engine", controller_source)
        self.assertIn("load_glossary_file(glossary_engine", translation_source)
        self.assertIn("load_glossary_file(glossary_engine", stress_source)
        self.assertNotIn("class GlossaryLoader", engine_source)
        self.assertNotIn("csv.reader", engine_source)
        self.assertNotIn("openpyxl", engine_source)

    def test_legacy_and_activated_sidecar_parity_for_identical_inputs(self) -> None:
        inputs = (
            "same",  # exact TM source that is also a glossary term (TM priority)
            "A high performance Glossary Engine",  # exact miss, glossary hit
            "Completely unrelated phrase",  # exact and glossary miss
        )

        def render_file_mode(tag: str) -> tuple[list[object], dict[str, int]]:
            input_path = root / f"input_{tag}.xlsx"
            output_path = root / f"output_{tag}.xlsx"
            timing_path = root / f"timing_{tag}.json"
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                self.fail("input workbook has no active sheet")
            for index, text in enumerate(inputs, start=1):
                sheet[f"A{index}"] = text
            workbook.save(input_path)

            _artifact_path = run_file_mode_benchmark(
                input_xlsx=input_path,
                output_xlsx=output_path,
                timing_json=timing_path,
                source_column="A",
                target_column="B",
                sheet_name=None,
                max_rows=None,
            )
            output = load_workbook(output_path, read_only=True)
            output_sheet = output.active
            if output_sheet is None:
                self.fail("output workbook has no active sheet")
            rendered = [
                cast(object, output_sheet[f"B{index}"].value)
                for index in range(1, len(inputs) + 1)
            ]
            output.close()
            status_counts = cast(
                dict[str, int],
                json.loads(timing_path.read_text(encoding="utf-8"))["status_counts"],
            )
            return rendered, status_counts

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tm_path = root / "tm.primary.jsonl"
            _ = tm_path.write_bytes(SOURCE_BYTES)
            glossary_path = root / "terms.csv"
            _ = glossary_path.write_text(
                "same,同様\nGlossary,术语\n",
                encoding="utf-8",
            )

            with patch.object(
                logic_controller, "TM_FILE", str(tm_path)
            ), patch.object(
                logic_controller, "GLOSSARY_FILE", str(glossary_path)
            ):
                legacy_controller = logic_controller.LogicController()
                legacy_results = [
                    legacy_controller.get_suggestions(text) for text in inputs
                ]
                legacy_rendered, legacy_counts = render_file_mode("legacy")

            identity, coordinator, _sealed, prepared, handle = _first_prepared(
                root,
                fts5_available=True,
            )
            self.assertEqual(
                identity.configured_jsonl_path.resolve(),
                tm_path.resolve(),
            )
            with patch("tm_sqlite_store._probe_fts5", return_value=True):
                _ = coordinator.publish_activation(prepared, handle)

            with patch.object(
                logic_controller, "TM_FILE", str(tm_path)
            ), patch.object(
                logic_controller, "GLOSSARY_FILE", str(glossary_path)
            ):
                activated_controller = logic_controller.LogicController()
                activated_results = [
                    activated_controller.get_suggestions(text) for text in inputs
                ]
                activated_rendered, activated_counts = render_file_mode("activated")

        self.assertEqual(activated_results, legacy_results)
        self.assertEqual(activated_rendered, legacy_rendered)
        self.assertEqual(activated_counts, legacy_counts)
        self.assertEqual(
            legacy_results[0],
            {
                "status": "TM_HIT",
                "tm_match": {
                    "source": "same",
                    "target": "winner",
                    "match_type": "EXACT",
                    "similarity": 1.0,
                },
            },
        )
        self.assertEqual(legacy_results[1]["status"], "TERMS_FOUND")
        terms = cast(list[dict[str, object]], legacy_results[1]["terms"])
        self.assertEqual(
            [(term["source_term"], term["target_term"]) for term in terms],
            [("Glossary", "术语")],
        )
        self.assertEqual(legacy_results[2], {"status": "NO_MATCH"})
        self.assertEqual(
            legacy_rendered,
            ["[TM] winner", "[Terms] Glossary->术语", "[No Match]"],
        )
        self.assertEqual(
            legacy_counts,
            {"TM_HIT": 1, "TERMS_FOUND": 1, "NO_MATCH": 1},
        )
        self.assertEqual(
            set(legacy_counts),
            {"TM_HIT", "TERMS_FOUND", "NO_MATCH"},
        )
        self.assertEqual(
            set(activated_counts),
            {"TM_HIT", "TERMS_FOUND", "NO_MATCH"},
        )
        self.assertEqual(
            {cast(str, result["status"]) for result in legacy_results},
            {"TM_HIT", "TERMS_FOUND", "NO_MATCH"},
        )
        self.assertEqual(
            {cast(str, result["status"]) for result in activated_results},
            {"TM_HIT", "TERMS_FOUND", "NO_MATCH"},
        )

    def test_logic_controller_self_test_matches_default_resources(self) -> None:
        result = subprocess.run(
            [sys.executable, str(PROJECT_ROOT / "logic_controller.py")],
            cwd=PROJECT_ROOT,
            check=False,
            capture_output=True,
            text=True,
        )

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertIn("Logic Controller Self-Test Passed", result.stdout)

    def test_headless_file_adapter_preserves_three_status_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.xlsx"
            output_path = root / "output.xlsx"
            timing_path = root / "timing.json"
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                self.fail("input workbook has no active sheet")
            sheet["A1"] = "System Ready"
            sheet["A2"] = "Initializing Glossary Engine"
            sheet["A3"] = "Completely unrelated phrase"
            workbook.save(input_path)

            result_path = run_file_mode_benchmark(
                input_xlsx=input_path,
                output_xlsx=output_path,
                timing_json=timing_path,
                source_column="A",
                target_column="B",
                sheet_name=None,
                max_rows=None,
            )
            output = load_workbook(output_path, read_only=True)
            output_sheet = output.active
            if output_sheet is None:
                self.fail("output workbook has no active sheet")
            values = [output_sheet[f"B{row}"].value for row in range(1, 4)]
            output.close()
            timing = json.loads(result_path.read_text(encoding="utf-8"))

        self.assertEqual(values[0], "[TM] 系统就绪")
        self.assertTrue(str(values[1]).startswith("[Terms]"))
        self.assertEqual(values[2], "[No Match]")
        self.assertEqual(
            timing["status_counts"],
            {"TM_HIT": 1, "TERMS_FOUND": 1, "NO_MATCH": 1},
        )

    def test_interactive_adapter_compiles_and_only_reaches_engine_via_logic(self) -> None:
        adapter_path = PROJECT_ROOT / "excel_adapter.py"
        _ = compile(
            adapter_path.read_text(encoding="utf-8"),
            str(adapter_path),
            "exec",
        )
        tree = ast.parse(adapter_path.read_text(encoding="utf-8"))
        imported_modules = {
            node.module
            for node in ast.walk(tree)
            if isinstance(node, ast.ImportFrom) and node.module
        }

        self.assertIn("logic_controller", imported_modules)
        self.assertNotIn("tm_engine", imported_modules)
        self.assertNotIn("glossary_engine", imported_modules)
        self.assertFalse(any(module.startswith("PySide6") for module in imported_modules))


if __name__ == "__main__":
    unittest.main()
