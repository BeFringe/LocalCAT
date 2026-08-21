from __future__ import annotations

import csv
from dataclasses import replace
import io
import json
from pathlib import Path
import tempfile
import threading
import time
import unittest
from unittest import mock


FIXTURE_ROOT = Path(__file__).parent / "fixtures" / "parser" / "termbase"


def _golden_cases() -> dict[str, dict[str, object]]:
    manifest = json.loads((FIXTURE_ROOT / "manifest.json").read_text(encoding="utf-8"))
    return {
        case["case_id"]: case
        for entry in manifest["formats"]
        for case in entry["cases"]
    }


def _golden_bytes(case_id: str) -> bytes:
    from tests.test_parser_termbase_golden import _case_bytes

    return _case_bytes(_golden_cases()[case_id])


def _selection(
    *,
    source: str | int,
    target: str | int,
    policy: str,
):
    from parser_contracts import (
        ColumnSelectorKind,
        TermbaseColumnSelection,
        TermbaseColumnSelector,
        TermbaseHeaderPolicy,
    )

    def selector(value: str | int) -> TermbaseColumnSelector:
        if type(value) is str:
            return TermbaseColumnSelector(
                kind=ColumnSelectorKind.HEADER_NAME,
                header_name=value,
            )
        return TermbaseColumnSelector(
            kind=ColumnSelectorKind.ZERO_BASED_INDEX,
            zero_based_index=value,
        )

    policies = {
        "first_row": TermbaseHeaderPolicy.FIRST_ROW,
        "no_header": TermbaseHeaderPolicy.NO_HEADER,
        "legacy_allowlist": TermbaseHeaderPolicy.LEGACY_ALLOWLIST,
    }
    return TermbaseColumnSelection(
        source=selector(source),
        target=selector(target),
        header_policy=policies[policy],
    )


class _ParserFixture(unittest.TestCase):
    def setUp(self) -> None:
        self._temporary = tempfile.TemporaryDirectory(prefix="termbase-codec-")
        self.root = Path(self._temporary.name)

    def tearDown(self) -> None:
        self._temporary.cleanup()

    def _input(self, name: str, payload: bytes):
        from parser_contracts import SourceReference

        path = self.root / name
        path.write_bytes(payload)
        return SourceReference(
            safe_root=str(self.root),
            selected_path=str(path),
            display_hint=name,
        )

    @staticmethod
    def _request(descriptor, columns):
        from parser_contracts import ReadRequest, TermbaseReadOptions

        return ReadRequest(
            purpose=descriptor.purpose,
            format_id=descriptor.format_id,
            termbase_options=TermbaseReadOptions(columns=columns),
        )

    def _materialize(self, *, descriptor, codec, name: str, payload: bytes, columns):
        from parser_source import create_sealed_snapshot, materialize

        snapshot = create_sealed_snapshot(
            self._input(name, payload),
            limit_profile=descriptor.limit_profile,
        )
        try:
            return materialize(codec, snapshot, self._request(descriptor, columns))
        finally:
            snapshot.close()

    def _validate(self, *, descriptor, codec, name: str, payload: bytes, columns):
        from parser_source import create_sealed_snapshot, validate

        snapshot = create_sealed_snapshot(
            self._input(name, payload),
            limit_profile=descriptor.limit_profile,
        )
        try:
            return validate(codec, snapshot, self._request(descriptor, columns))
        finally:
            snapshot.close()


class TermbaseCodecContractTests(unittest.TestCase):
    def test_descriptors_publish_distinct_frozen_term_profiles(self) -> None:
        from parser_contracts import (
            EffectivePurpose,
            InputConsumptionPolicy,
            TERMBASE_CSV_V1,
            TERMBASE_XLSX_V1,
        )
        from parser_termbase_codec import (
            TERMBASE_CSV_DESCRIPTOR,
            TERMBASE_XLSX_DESCRIPTOR,
        )

        csv_descriptor = TERMBASE_CSV_DESCRIPTOR
        xlsx_descriptor = TERMBASE_XLSX_DESCRIPTOR
        self.assertIs(csv_descriptor.purpose, EffectivePurpose.TERMBASE)
        self.assertEqual(csv_descriptor.format_id, TERMBASE_CSV_V1)
        self.assertIs(
            csv_descriptor.input_consumption_policy,
            InputConsumptionPolicy.SEALED_BYTES_EOF,
        )
        self.assertEqual(xlsx_descriptor.format_id, TERMBASE_XLSX_V1)
        self.assertIs(
            xlsx_descriptor.input_consumption_policy,
            InputConsumptionPolicy.XLSX_PREFLIGHT_ACTIVE_SHEET,
        )
        self.assertFalse(csv_descriptor.capabilities.canonical_write)
        self.assertFalse(xlsx_descriptor.capabilities.canonical_write)
        self.assertFalse(csv_descriptor.capabilities.active_sheet_only)
        self.assertTrue(xlsx_descriptor.capabilities.active_sheet_only)
        self.assertIsNot(csv_descriptor.limit_profile, xlsx_descriptor.limit_profile)

        self.assertEqual(csv_descriptor.limit_profile.max_input_bytes, 100 * 1024 * 1024)
        self.assertEqual(csv_descriptor.limit_profile.max_records, 1_000_000)
        self.assertEqual(xlsx_descriptor.limit_profile.max_records, 1_048_576)
        self.assertEqual(xlsx_descriptor.limit_profile.max_archive_members, 4_096)
        self.assertEqual(xlsx_descriptor.limit_profile.max_expanded_bytes, 256 * 1024 * 1024)
        self.assertEqual(xlsx_descriptor.limit_profile.max_compression_ratio, 100.0)
        for descriptor in (csv_descriptor, xlsx_descriptor):
            self.assertTrue(descriptor.capabilities.readable)
            self.assertTrue(descriptor.capabilities.validatable)
            self.assertFalse(descriptor.capabilities.source_round_trip_write)
            self.assertIsNone(descriptor.canonical_serializer_factory)
            self.assertEqual(
                descriptor.declared_issue_codes,
                tuple(sorted(set(descriptor.declared_issue_codes))),
            )
            self.assertIn(
                "PARSER.TERMBASE.HEADER_SKIPPED",
                descriptor.declared_issue_codes,
            )


class CsvTermbaseCodecTests(_ParserFixture):
    @staticmethod
    def _csv_descriptor(*, max_field_chars: int, max_input_bytes: int):
        from parser_termbase_codec import TERMBASE_CSV_DESCRIPTOR

        profile_id = f"termbase-csv-field-{max_field_chars}-test"
        return replace(
            TERMBASE_CSV_DESCRIPTOR,
            limit_profile=replace(
                TERMBASE_CSV_DESCRIPTOR.limit_profile,
                profile_id=profile_id,
                max_input_bytes=max_input_bytes,
                max_decoded_field_chars=max_field_chars,
            ),
            capabilities=replace(
                TERMBASE_CSV_DESCRIPTOR.capabilities,
                format_profile=profile_id,
            ),
        )

    def test_header_index_legacy_and_bom_golden_inputs(self) -> None:
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        cases = (
            (
                "csv-header-name-valid",
                _selection(source="Source", target="Target", policy="first_row"),
                (("row-2", "Alpha", "甲"), ("row-3", "Beta", "乙")),
                (("PARSER.TERMBASE.HEADER_SKIPPED", 1),),
            ),
            (
                "csv-index-headerless-valid",
                _selection(source=1, target=0, policy="no_header"),
                (("row-1", "Alpha", "甲"), ("row-2", "Beta", "乙")),
                (),
            ),
            (
                "csv-legacy-preset-valid",
                _selection(source=0, target=1, policy="legacy_allowlist"),
                (("row-2", "Alpha", "甲"), ("row-3", "Beta", "乙")),
                (("PARSER.TERMBASE.HEADER_SKIPPED", 1),),
            ),
            (
                "csv-utf8-bom-valid",
                _selection(source="Source", target="Target", policy="first_row"),
                (("row-2", "Alpha", "甲"),),
                (("PARSER.TERMBASE.HEADER_SKIPPED", 1),),
            ),
        )
        for case_id, columns, expected, expected_issues in cases:
            with self.subTest(case_id=case_id):
                result = self._materialize(
                    descriptor=TERMBASE_CSV_DESCRIPTOR,
                    codec=CsvTermbaseCodec(),
                    name=f"{case_id}.csv",
                    payload=_golden_bytes(case_id),
                    columns=columns,
                )
                observed = tuple(
                    (record.local_id, record.source, record.target)
                    for record in result.records
                )
                self.assertEqual(observed, expected)
                self.assertEqual(
                    tuple((issue.code, issue.record_number) for issue in result.issues),
                    expected_issues,
                )

    def test_first_row_header_skip_warning_is_body_safe_and_nonfatal(self) -> None:
        from parser_contracts import IssueSeverity
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        source_header = "PRIVATE_SOURCE_HEADER"
        target_header = "PRIVATE_TARGET_HEADER"
        result = self._materialize(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="body-safe-header.csv",
            payload=f"{source_header},{target_header}\nSource,Target\n".encode(),
            columns=_selection(
                source=source_header,
                target=target_header,
                policy="first_row",
            ),
        )

        self.assertEqual(tuple(record.local_id for record in result.records), ("row-2",))
        self.assertEqual(result.terminal.record_count, 1)
        self.assertEqual(len(result.issues), 1)
        issue = result.issues[0]
        self.assertIs(issue.severity, IssueSeverity.WARNING)
        self.assertEqual(issue.code, "PARSER.TERMBASE.HEADER_SKIPPED")
        self.assertEqual(issue.record_number, 1)
        self.assertNotIn(source_header, issue.safe_summary)
        self.assertNotIn(target_header, issue.safe_summary)

    def test_header_matching_is_trimmed_exact_case_sensitive_and_index_first_row_skips(self) -> None:
        from parser_contracts import ValidationOutcome
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        payload = "  Source  , Target \n  A  B  ,  甲  \n".encode()
        result = self._materialize(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="trim.csv",
            payload=payload,
            columns=_selection(source="Source", target="Target", policy="first_row"),
        )
        self.assertEqual(
            tuple((row.local_id, row.source, row.target) for row in result.records),
            (("row-2", "A  B", "甲"),),
        )

        mismatched = self._validate(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="case.csv",
            payload=payload,
            columns=_selection(source="source", target="Target", policy="first_row"),
        )
        self.assertIs(mismatched.outcome, ValidationOutcome.FAILED)
        self.assertEqual(mismatched.issues[-1].code, "PARSER.TERMBASE.HEADER_MISSING")

        index_result = self._materialize(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="index-header.csv",
            payload=b"Ignored,Ignored\nTarget,Source\n",
            columns=_selection(source=1, target=0, policy="first_row"),
        )
        self.assertEqual(
            tuple((row.local_id, row.source, row.target) for row in index_result.records),
            (("row-2", "Source", "Target"),),
        )

    def test_warning_rows_keep_holes_order_and_duplicate_sources(self) -> None:
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        result = self._materialize(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="warnings.csv",
            payload=_golden_bytes("csv-record-warnings") + "Alpha,终值\n".encode(),
            columns=_selection(source="Source", target="Target", policy="first_row"),
        )
        self.assertEqual(
            tuple(record.local_id for record in result.records),
            ("row-2", "row-7", "row-8"),
        )
        self.assertEqual(
            tuple(record.source for record in result.records),
            ("Alpha", "Beta", "Alpha"),
        )
        self.assertEqual(
            tuple(issue.record_number for issue in result.issues),
            (1, 3, 4, 5, 6),
        )
        self.assertEqual(
            tuple(issue.code for issue in result.issues),
            (
                "PARSER.TERMBASE.HEADER_SKIPPED",
                "PARSER.TERMBASE.ROW_EMPTY",
                "PARSER.TERMBASE.ROW_MISSING_COLUMN",
                "PARSER.TERMBASE.SOURCE_EMPTY",
                "PARSER.TERMBASE.TARGET_EMPTY",
            ),
        )

    def test_semantically_empty_csv_rows_are_distinct_from_one_sided_rows(self) -> None:
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        result = self._materialize(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="empty-row-matrix.csv",
            payload=(
                b"\n"
                b",\n"
                b"  ,  \n"
                b"  , \t ,   \n"
                b",Target\n"
                b"Source,\n"
                b"Accepted,Translation\n"
            ),
            columns=_selection(source=0, target=1, policy="no_header"),
        )

        self.assertEqual(
            tuple((record.local_id, record.source, record.target) for record in result.records),
            (("row-7", "Accepted", "Translation"),),
        )
        self.assertEqual(
            tuple((issue.code, issue.record_number) for issue in result.issues),
            (
                ("PARSER.TERMBASE.ROW_EMPTY", 1),
                ("PARSER.TERMBASE.ROW_EMPTY", 2),
                ("PARSER.TERMBASE.ROW_EMPTY", 3),
                ("PARSER.TERMBASE.ROW_EMPTY", 4),
                ("PARSER.TERMBASE.SOURCE_EMPTY", 5),
                ("PARSER.TERMBASE.TARGET_EMPTY", 6),
            ),
        )

    def test_warning_mapping_can_stage_without_mutating_managed_target(self) -> None:
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        managed_target = self.root / "managed-target.csv"
        managed_bytes = "Keep,保留\n".encode()
        managed_target.write_bytes(managed_bytes)
        result = self._materialize(
            descriptor=TERMBASE_CSV_DESCRIPTOR,
            codec=CsvTermbaseCodec(),
            name="adjacent-import.csv",
            payload=b"Source,Target\n,\nNew,Value\n",
            columns=_selection(source=0, target=1, policy="legacy_allowlist"),
        )

        skipped_codes = {
            "PARSER.TERMBASE.HEADER_SKIPPED",
            "PARSER.TERMBASE.ROW_EMPTY",
            "PARSER.TERMBASE.ROW_MISSING_COLUMN",
            "PARSER.TERMBASE.SOURCE_EMPTY",
            "PARSER.TERMBASE.TARGET_EMPTY",
        }
        skipped = sum(issue.code in skipped_codes for issue in result.issues)
        staged = {"Keep": "保留"}
        staged.update((record.source, record.target) for record in result.records)

        self.assertEqual(skipped, 2)
        self.assertEqual(
            tuple(issue.code for issue in result.issues),
            (
                "PARSER.TERMBASE.HEADER_SKIPPED",
                "PARSER.TERMBASE.ROW_EMPTY",
            ),
        )
        self.assertEqual(staged, {"Keep": "保留", "New": "Value"})
        self.assertEqual(managed_target.read_bytes(), managed_bytes)

    def test_header_failures_and_mixed_selector_same_column_are_pre_record_fatal(self) -> None:
        from parser_contracts import ValidationOutcome
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        cases = (
            (
                "csv-missing-header",
                _selection(source="Source", target="Target", policy="first_row"),
                "PARSER.TERMBASE.HEADER_MISSING",
            ),
            (
                "csv-duplicate-header",
                _selection(source="Source", target="Target", policy="first_row"),
                "PARSER.TERMBASE.HEADER_DUPLICATE",
            ),
            (
                "csv-header-name-valid",
                _selection(source="Source", target=0, policy="first_row"),
                "PARSER.TERMBASE.COLUMN_SELECTION_SAME",
            ),
        )
        for case_id, columns, code in cases:
            with self.subTest(case_id=case_id):
                report = self._validate(
                    descriptor=TERMBASE_CSV_DESCRIPTOR,
                    codec=CsvTermbaseCodec(),
                    name=f"{case_id}.csv",
                    payload=_golden_bytes(case_id),
                    columns=columns,
                )
                self.assertIs(report.outcome, ValidationOutcome.FAILED)
                self.assertEqual(report.provisional_record_count, 0)
                self.assertEqual(report.issues[-1].code, code)
                self.assertIsNone(report.terminal)

    def test_encoding_and_fatal_tail_have_stable_fatal_mapping(self) -> None:
        from parser_contracts import ValidationOutcome
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        cases = (
            ("csv-invalid-utf8", "PARSER.SOURCE.ENCODING_FAILED", 0),
            ("csv-fatal-tail", "PARSER.SYNTAX.MALFORMED", 1),
        )
        columns = _selection(source="Source", target="Target", policy="first_row")
        for case_id, code, provisional in cases:
            with self.subTest(case_id=case_id):
                report = self._validate(
                    descriptor=TERMBASE_CSV_DESCRIPTOR,
                    codec=CsvTermbaseCodec(),
                    name=f"{case_id}.csv",
                    payload=_golden_bytes(case_id),
                    columns=columns,
                )
                self.assertIs(report.outcome, ValidationOutcome.FAILED)
                self.assertEqual(report.provisional_record_count, provisional)
                self.assertEqual(report.issues[-1].code, code)

    def test_cancellation_after_first_row_denies_terminal(self) -> None:
        from parser_source import (
            CancellationToken,
            GuardedParseSession,
            ParserSessionError,
            create_sealed_snapshot,
        )
        from parser_termbase_codec import CsvTermbaseCodec, TERMBASE_CSV_DESCRIPTOR

        token = CancellationToken()
        snapshot = create_sealed_snapshot(
            self._input("cancel.csv", _golden_bytes("csv-cancel-after-row")),
            limit_profile=TERMBASE_CSV_DESCRIPTOR.limit_profile,
        )
        session = GuardedParseSession(
            CsvTermbaseCodec(),
            snapshot,
            self._request(
                TERMBASE_CSV_DESCRIPTOR,
                _selection(source=0, target=1, policy="no_header"),
            ),
            cancellation=token,
        )
        iterator = iter(session)
        try:
            self.assertEqual(next(iterator).local_id, "row-1")
            token.cancel()
            with self.assertRaises(ParserSessionError) as caught:
                next(iterator)
            self.assertEqual(caught.exception.code, "PARSER.SOURCE.CANCELLED")
            with self.assertRaises(ParserSessionError):
                session.verified_terminal()
        finally:
            iterator.close()
            session.close()
            snapshot.close()

    def test_descriptor_field_limit_overrides_stdlib_default_and_restores_global(self) -> None:
        from parser_contracts import ValidationOutcome
        from parser_termbase_codec import CsvTermbaseCodec

        descriptor = self._csv_descriptor(
            max_field_chars=131_073,
            max_input_bytes=200_000,
        )
        columns = _selection(source=0, target=1, policy="no_header")
        original = csv.field_size_limit()
        sentinel = 65_537
        csv.field_size_limit(sentinel)
        try:
            for length in (131_072, 131_073):
                with self.subTest(length=length):
                    result = self._materialize(
                        descriptor=descriptor,
                        codec=CsvTermbaseCodec(descriptor),
                        name=f"field-{length}.csv",
                        payload=("S" * length + ",T\n").encode(),
                        columns=columns,
                    )
                    self.assertEqual(len(result.records[0].source), length)
                    self.assertEqual(csv.field_size_limit(), sentinel)

            report = self._validate(
                descriptor=descriptor,
                codec=CsvTermbaseCodec(descriptor),
                name="field-over.csv",
                payload=("S" * 131_074 + ",T\n").encode(),
                columns=columns,
            )
            self.assertIs(report.outcome, ValidationOutcome.FAILED)
            self.assertEqual(report.issues[-1].code, "PARSER.LIMIT.FIELD")
            self.assertEqual(csv.field_size_limit(), sentinel)

            malformed = self._validate(
                descriptor=descriptor,
                codec=CsvTermbaseCodec(descriptor),
                name="field-malformed.csv",
                payload=b'Good,T\n"unterminated\n',
                columns=columns,
            )
            self.assertIs(malformed.outcome, ValidationOutcome.FAILED)
            self.assertEqual(malformed.issues[-1].code, "PARSER.SYNTAX.MALFORMED")
            self.assertEqual(csv.field_size_limit(), sentinel)
        finally:
            csv.field_size_limit(original)

    def test_global_field_limit_is_restored_before_yield_cancel_and_early_close(self) -> None:
        from parser_source import (
            CancellationToken,
            GuardedParseSession,
            ParserSessionError,
            create_sealed_snapshot,
        )
        from parser_termbase_codec import CsvTermbaseCodec

        descriptor = self._csv_descriptor(
            max_field_chars=150_000,
            max_input_bytes=200_000,
        )
        columns = _selection(source=0, target=1, policy="no_header")
        original = csv.field_size_limit()
        sentinel = 65_539
        csv.field_size_limit(sentinel)
        snapshot = create_sealed_snapshot(
            self._input("restore.csv", b"First,T\nSecond,U\n"),
            limit_profile=descriptor.limit_profile,
        )
        token = CancellationToken()
        session = GuardedParseSession(
            CsvTermbaseCodec(descriptor),
            snapshot,
            self._request(descriptor, columns),
            cancellation=token,
        )
        iterator = iter(session)
        try:
            self.assertEqual(next(iterator).local_id, "row-1")
            self.assertEqual(csv.field_size_limit(), sentinel)
            token.cancel()
            with self.assertRaises(ParserSessionError) as caught:
                next(iterator)
            self.assertEqual(caught.exception.code, "PARSER.SOURCE.CANCELLED")
            self.assertEqual(csv.field_size_limit(), sentinel)
        finally:
            iterator.close()
            session.close()
            snapshot.close()
            self.assertEqual(csv.field_size_limit(), sentinel)
            csv.field_size_limit(original)

        csv.field_size_limit(sentinel)
        snapshot = create_sealed_snapshot(
            self._input("early-close.csv", b"First,T\nSecond,U\n"),
            limit_profile=descriptor.limit_profile,
        )
        session = GuardedParseSession(
            CsvTermbaseCodec(descriptor),
            snapshot,
            self._request(descriptor, columns),
        )
        iterator = iter(session)
        try:
            self.assertEqual(next(iterator).local_id, "row-1")
            self.assertEqual(csv.field_size_limit(), sentinel)
        finally:
            iterator.close()
            session.close()
            snapshot.close()
            self.assertEqual(csv.field_size_limit(), sentinel)
            csv.field_size_limit(original)

    def test_concurrent_csv_profiles_do_not_cross_talk_global_field_limit(self) -> None:
        from parser_termbase_codec import CsvTermbaseCodec

        columns = _selection(source=0, target=1, policy="no_header")
        descriptors = (
            self._csv_descriptor(max_field_chars=140_000, max_input_bytes=180_000),
            self._csv_descriptor(max_field_chars=150_000, max_input_bytes=190_000),
        )
        lengths = (135_000, 145_000)
        original = csv.field_size_limit()
        sentinel = 65_541
        csv.field_size_limit(sentinel)
        real_field_size_limit = csv.field_size_limit
        observation_lock = threading.Lock()
        active_temporary_values = 0
        maximum_active = 0

        def observed_field_size_limit(value=None):
            nonlocal active_temporary_values, maximum_active
            if value is None:
                return real_field_size_limit()
            previous = real_field_size_limit(value)
            with observation_lock:
                if value == sentinel:
                    active_temporary_values -= 1
                else:
                    active_temporary_values += 1
                    maximum_active = max(maximum_active, active_temporary_values)
            if value != sentinel:
                time.sleep(0.005)
            return previous

        barrier = threading.Barrier(2)
        results: list[int] = []
        errors: list[BaseException] = []

        def worker(index: int) -> None:
            try:
                barrier.wait()
                result = self._materialize(
                    descriptor=descriptors[index],
                    codec=CsvTermbaseCodec(descriptors[index]),
                    name=f"concurrent-{index}.csv",
                    payload=("S" * lengths[index] + ",T\n").encode(),
                    columns=columns,
                )
                results.append(len(result.records[0].source))
            except BaseException as exc:
                errors.append(exc)

        try:
            with mock.patch(
                "parser_termbase_codec.csv.field_size_limit",
                side_effect=observed_field_size_limit,
            ):
                threads = tuple(
                    threading.Thread(target=worker, args=(index,))
                    for index in range(2)
                )
                for thread in threads:
                    thread.start()
                for thread in threads:
                    thread.join(timeout=5)
                self.assertTrue(all(not thread.is_alive() for thread in threads))
            self.assertEqual(errors, [])
            self.assertEqual(sorted(results), sorted(lengths))
            self.assertEqual(maximum_active, 1)
            self.assertEqual(active_temporary_values, 0)
            self.assertEqual(real_field_size_limit(), sentinel)
        finally:
            real_field_size_limit(original)


class XlsxTermbaseCodecTests(_ParserFixture):
    def test_real_preflight_projects_descriptor_limits_and_active_sheet_only(self) -> None:
        import parser_termbase_codec as codec_module
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        actual_preflight = codec_module.preflight_xlsx
        observed_limits = []

        def checked_preflight(source, limits):
            observed_limits.append(limits)
            return actual_preflight(source, limits)

        with mock.patch.object(codec_module, "preflight_xlsx", side_effect=checked_preflight):
            result = self._materialize(
                descriptor=TERMBASE_XLSX_DESCRIPTOR,
                codec=XlsxTermbaseCodec(),
                name="active.xlsx",
                payload=_golden_bytes("xlsx-header-name-valid"),
                columns=_selection(source="Source", target="Target", policy="first_row"),
            )

        self.assertEqual(
            tuple((record.local_id, record.source) for record in result.records),
            (("row-2", "Alpha"), ("row-3", "Beta")),
        )
        self.assertEqual(
            tuple((issue.code, issue.record_number) for issue in result.issues),
            (("PARSER.TERMBASE.HEADER_SKIPPED", 1),),
        )
        self.assertEqual(len(observed_limits), 1)
        limits = observed_limits[0]
        profile = TERMBASE_XLSX_DESCRIPTOR.limit_profile
        self.assertEqual(limits.max_archive_members, profile.max_archive_members)
        self.assertEqual(limits.max_expanded_bytes, profile.max_expanded_bytes)
        self.assertEqual(limits.max_compression_ratio, profile.max_compression_ratio)
        self.assertEqual(limits.max_xml_depth, profile.max_structure_depth)

    def test_xlsx_preserves_duplicate_source_rows_without_non_active_aggregation(self) -> None:
        import openpyxl
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        workbook = openpyxl.Workbook()
        active = workbook.active
        active.title = "Active"
        active.append(["Source", "Target"])
        active.append(["Same", "甲"])
        active.append(["Same", "乙"])
        ignored = workbook.create_sheet("Ignored")
        ignored.append(["Source", "Target"])
        ignored.append(["DO_NOT_AGGREGATE", "忽略"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = self._materialize(
            descriptor=TERMBASE_XLSX_DESCRIPTOR,
            codec=XlsxTermbaseCodec(),
            name="duplicates.xlsx",
            payload=buffer.getvalue(),
            columns=_selection(source="Source", target="Target", policy="first_row"),
        )
        self.assertEqual(
            tuple((row.local_id, row.source, row.target) for row in result.records),
            (("row-2", "Same", "甲"), ("row-3", "Same", "乙")),
        )

    def test_xlsx_reuses_index_legacy_warning_and_duplicate_row_semantics(self) -> None:
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        for case_id, columns, expected_ids, expected_issues in (
            (
                "xlsx-index-headerless-valid",
                _selection(source=1, target=0, policy="no_header"),
                ("row-1", "row-2"),
                (),
            ),
            (
                "xlsx-legacy-preset-valid",
                _selection(source=0, target=1, policy="legacy_allowlist"),
                ("row-2", "row-3"),
                (("PARSER.TERMBASE.HEADER_SKIPPED", 1),),
            ),
        ):
            with self.subTest(case_id=case_id):
                result = self._materialize(
                    descriptor=TERMBASE_XLSX_DESCRIPTOR,
                    codec=XlsxTermbaseCodec(),
                    name=f"{case_id}.xlsx",
                    payload=_golden_bytes(case_id),
                    columns=columns,
                )
                self.assertEqual(
                    tuple(record.local_id for record in result.records),
                    expected_ids,
                )
                self.assertEqual(
                    tuple((issue.code, issue.record_number) for issue in result.issues),
                    expected_issues,
                )

        warning_result = self._materialize(
            descriptor=TERMBASE_XLSX_DESCRIPTOR,
            codec=XlsxTermbaseCodec(),
            name="warnings.xlsx",
            payload=_golden_bytes("xlsx-record-warnings"),
            columns=_selection(source="Source", target="Target", policy="first_row"),
        )
        self.assertEqual(
            tuple(record.local_id for record in warning_result.records),
            ("row-2", "row-7"),
        )
        self.assertEqual(
            tuple(issue.record_number for issue in warning_result.issues),
            (1, 3, 4, 5, 6),
        )
        self.assertEqual(
            tuple(issue.code for issue in warning_result.issues),
            (
                "PARSER.TERMBASE.HEADER_SKIPPED",
                "PARSER.TERMBASE.ROW_EMPTY",
                "PARSER.TERMBASE.ROW_MISSING_COLUMN",
                "PARSER.TERMBASE.SOURCE_EMPTY",
                "PARSER.TERMBASE.TARGET_EMPTY",
            ),
        )

    def test_semantically_empty_xlsx_rows_are_distinct_from_one_sided_rows(self) -> None:
        import openpyxl
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([None, None])
        sheet.append(["", ""])
        sheet.append(["  ", "\t", "   "])
        sheet.append(["", "Target"])
        sheet.append(["Source", ""])
        sheet.append(["Accepted", "Translation"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        result = self._materialize(
            descriptor=TERMBASE_XLSX_DESCRIPTOR,
            codec=XlsxTermbaseCodec(),
            name="empty-row-matrix.xlsx",
            payload=buffer.getvalue(),
            columns=_selection(source=0, target=1, policy="no_header"),
        )

        self.assertEqual(
            tuple((record.local_id, record.source, record.target) for record in result.records),
            (("row-6", "Accepted", "Translation"),),
        )
        self.assertEqual(
            tuple((issue.code, issue.record_number) for issue in result.issues),
            (
                ("PARSER.TERMBASE.ROW_EMPTY", 1),
                ("PARSER.TERMBASE.ROW_EMPTY", 2),
                ("PARSER.TERMBASE.ROW_EMPTY", 3),
                ("PARSER.TERMBASE.SOURCE_EMPTY", 4),
                ("PARSER.TERMBASE.TARGET_EMPTY", 5),
            ),
        )

    def test_xlsx_header_failures_are_pre_record_fatal(self) -> None:
        from parser_contracts import ValidationOutcome
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        cases = (
            (
                "xlsx-missing-header",
                _selection(source="Source", target="Target", policy="first_row"),
                "PARSER.TERMBASE.HEADER_MISSING",
            ),
            (
                "xlsx-duplicate-header",
                _selection(source="Source", target="Target", policy="first_row"),
                "PARSER.TERMBASE.HEADER_DUPLICATE",
            ),
            (
                "xlsx-header-name-valid",
                _selection(source="Source", target=0, policy="first_row"),
                "PARSER.TERMBASE.COLUMN_SELECTION_SAME",
            ),
        )
        for case_id, columns, code in cases:
            with self.subTest(case_id=case_id):
                report = self._validate(
                    descriptor=TERMBASE_XLSX_DESCRIPTOR,
                    codec=XlsxTermbaseCodec(),
                    name=f"{case_id}.xlsx",
                    payload=_golden_bytes(case_id),
                    columns=columns,
                )
                self.assertIs(report.outcome, ValidationOutcome.FAILED)
                self.assertEqual(report.provisional_record_count, 0)
                self.assertEqual(report.issues[-1].code, code)

    def test_xlsx_preflight_danger_and_limits_fail_before_raw_events(self) -> None:
        from parser_contracts import ValidationOutcome
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        cases = (
            ("xlsx-dtd-member", "PARSER.TERMBASE.XML_DECLARATION_FORBIDDEN"),
            ("xlsx-entity-member", "PARSER.TERMBASE.XML_DECLARATION_FORBIDDEN"),
            ("xlsx-invalid-xml-encoding", "PARSER.SOURCE.ENCODING_FAILED"),
            ("xlsx-archive-ratio-limit", "PARSER.LIMIT.COMPRESSION_RATIO"),
            ("xlsx-fatal-tail", "PARSER.SYNTAX.MALFORMED"),
        )
        columns = _selection(source="Source", target="Target", policy="first_row")
        for case_id, code in cases:
            with self.subTest(case_id=case_id):
                report = self._validate(
                    descriptor=TERMBASE_XLSX_DESCRIPTOR,
                    codec=XlsxTermbaseCodec(),
                    name=f"{case_id}.xlsx",
                    payload=_golden_bytes(case_id),
                    columns=columns,
                )
                self.assertIs(report.outcome, ValidationOutcome.FAILED)
                self.assertEqual(report.provisional_record_count, 0)
                self.assertEqual(report.issues[-1].code, code)
                self.assertIsNone(report.terminal)

    def test_openpyxl_is_conditional_and_receives_nonexecuting_options(self) -> None:
        import parser_termbase_codec as codec_module
        import openpyxl
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        calls = []
        actual_load = openpyxl.load_workbook

        def checked_load(*args, **kwargs):
            calls.append(kwargs.copy())
            return actual_load(*args, **kwargs)

        with mock.patch.object(openpyxl, "load_workbook", side_effect=checked_load):
            self._materialize(
                descriptor=TERMBASE_XLSX_DESCRIPTOR,
                codec=XlsxTermbaseCodec(),
                name="options.xlsx",
                payload=_golden_bytes("xlsx-header-name-valid"),
                columns=_selection(source="Source", target="Target", policy="first_row"),
            )
        self.assertEqual(
            calls,
            [
                {
                    "read_only": True,
                    "data_only": True,
                    "keep_links": False,
                    "keep_vba": False,
                }
            ],
        )

        with mock.patch.object(
            codec_module.importlib,
            "import_module",
            side_effect=ImportError("missing"),
        ):
            report = self._validate(
                descriptor=TERMBASE_XLSX_DESCRIPTOR,
                codec=XlsxTermbaseCodec(),
                name="missing.xlsx",
                payload=_golden_bytes("xlsx-openpyxl-missing"),
                columns=_selection(source="Source", target="Target", policy="first_row"),
            )
        self.assertEqual(
            report.issues[-1].code,
            "PARSER.CAPABILITY.CONDITIONAL_DEPENDENCY_MISSING",
        )

    def test_xlsx_cancellation_after_first_row_denies_terminal(self) -> None:
        from parser_source import (
            CancellationToken,
            GuardedParseSession,
            ParserSessionError,
            create_sealed_snapshot,
        )
        from parser_termbase_codec import XlsxTermbaseCodec, TERMBASE_XLSX_DESCRIPTOR

        token = CancellationToken()
        snapshot = create_sealed_snapshot(
            self._input("cancel.xlsx", _golden_bytes("xlsx-cancel-after-row")),
            limit_profile=TERMBASE_XLSX_DESCRIPTOR.limit_profile,
        )
        session = GuardedParseSession(
            XlsxTermbaseCodec(),
            snapshot,
            self._request(
                TERMBASE_XLSX_DESCRIPTOR,
                _selection(source=0, target=1, policy="no_header"),
            ),
            cancellation=token,
        )
        iterator = iter(session)
        try:
            self.assertEqual(next(iterator).local_id, "row-1")
            token.cancel()
            with self.assertRaises(ParserSessionError) as caught:
                next(iterator)
            self.assertEqual(caught.exception.code, "PARSER.SOURCE.CANCELLED")
            with self.assertRaises(ParserSessionError):
                session.verified_terminal()
        finally:
            iterator.close()
            session.close()
            snapshot.close()


if __name__ == "__main__":
    unittest.main()
