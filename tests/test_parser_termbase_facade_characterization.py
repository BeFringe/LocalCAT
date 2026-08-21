"""Characterization coverage for the legacy termbase import facades.

These tests freeze the observable CSV/XLSX behavior that the Parser
rebaseline must preserve while row selection moves behind a codec boundary.
They intentionally exercise the current public facades rather than a future
Parser contract.
"""

from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from editor_contracts import LegacyTermRow
from resource_importer import (
    ImportFailure,
    import_termbase,
    read_legacy_termbase_import,
)


class ParserTermbaseFacadeCharacterizationTests(unittest.TestCase):
    @staticmethod
    def _read_csv(path: Path) -> list[list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.reader(handle))

    def test_read_facade_is_side_effect_free_and_preserves_rows_and_ordinals(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "incoming.csv"
            sentinel = root / "managed.csv"
            source.write_text(
                "Source,Target,Notes\n"
                " Alpha , 甲 ,first\n"
                ",\n"
                "Alpha,乙,duplicate\n"
                "incomplete\n"
                "Beta,丙,final\n",
                encoding="utf-8-sig",
            )
            sentinel.write_bytes(b"managed-bytes-must-not-change")
            source_before = source.read_bytes()
            sentinel_before = sentinel.read_bytes()
            names_before = {path.name for path in root.iterdir()}

            with patch(
                "resource_importer._atomic_write_text",
                side_effect=AssertionError("read facade attempted a write"),
            ):
                rows, skipped = read_legacy_termbase_import(source)

            names_after = {path.name for path in root.iterdir()}
            self.assertEqual(source.read_bytes(), source_before)
            self.assertEqual(sentinel.read_bytes(), sentinel_before)

        self.assertIsInstance(rows, tuple)
        self.assertEqual(
            rows,
            (
                LegacyTermRow("Alpha", "甲", 1),
                LegacyTermRow("Alpha", "乙", 3),
                LegacyTermRow("Beta", "丙", 5),
            ),
        )
        self.assertEqual(skipped, 3)
        self.assertEqual(names_after, names_before)

    def test_read_facade_raises_import_failure_when_no_valid_row_exists(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "empty.csv"
            source.write_text(
                "SOURCE TEXT,TRANSLATION\n"
                ",\n"
                "only-one-column\n",
                encoding="utf-8-sig",
            )
            original = source.read_bytes()

            with self.assertRaisesRegex(
                ImportFailure,
                "no valid source/target rows",
            ):
                _ = read_legacy_termbase_import(source)

            self.assertEqual(source.read_bytes(), original)

    def test_transaction_facade_keeps_legacy_header_skip_and_source_lww(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "incoming.csv"
            target = root / "managed.csv"
            source.write_text(
                "SOURCE TEXT,TRANSLATION,Ignored\n"
                "Alpha,first,one\n"
                ",missing-source,two\n"
                "short-row\n"
                "Alpha,last,three\n"
                "Beta,value,four\n",
                encoding="utf-8-sig",
            )
            target.write_text(
                "Alpha,old\nKeep,stable\n",
                encoding="utf-8-sig",
            )

            report = import_termbase(source, target)
            rendered = self._read_csv(target)

        self.assertEqual(
            (
                report.imported,
                report.skipped,
                report.overwritten,
                report.errors,
            ),
            (2, 3, 2, ()),
        )
        self.assertEqual(
            rendered,
            [
                ["Alpha", "last"],
                ["Keep", "stable"],
                ["Beta", "value"],
            ],
        )

    def test_transaction_failure_preserves_target_and_removes_temp_file(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "incoming.csv"
            target = root / "managed.csv"
            source.write_text("Fresh,new\n", encoding="utf-8-sig")
            target.write_bytes(b"\xef\xbb\xbfKeep,stable\r\n")
            original = target.read_bytes()

            with patch(
                "resource_importer.os.replace",
                side_effect=OSError("injected replace failure"),
            ):
                report = import_termbase(source, target)

            remaining = {path.name for path in root.iterdir()}
            self.assertEqual(target.read_bytes(), original)

        self.assertTrue(report.errors)
        self.assertIn("injected replace failure", report.errors[0])
        self.assertEqual(remaining, {source.name, target.name})

    def test_xlsx_facades_read_only_the_active_worksheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "multi-sheet.xlsx"
            target = root / "managed.csv"
            workbook = Workbook()
            inactive = workbook.active
            self.assertIsNotNone(inactive)
            assert inactive is not None
            inactive.title = "Inactive"
            inactive.append(["Source", "Target"])
            inactive.append(["Wrong", "must-not-import"])
            active = workbook.create_sheet("Selected")
            active.append(["Source term", "Target term"])
            active.append(["Chosen", "selected-value"])
            trailing = workbook.create_sheet("Trailing")
            trailing.append(["Also wrong", "must-not-import"])
            workbook.active = workbook.sheetnames.index("Selected")
            workbook.save(source)
            workbook.close()

            rows, skipped = read_legacy_termbase_import(source)
            report = import_termbase(source, target)
            rendered = self._read_csv(target)

        self.assertEqual(rows, (LegacyTermRow("Chosen", "selected-value", 1),))
        self.assertEqual(skipped, 1)
        self.assertEqual(
            (report.imported, report.skipped, report.overwritten, report.errors),
            (1, 1, 0, ()),
        )
        self.assertEqual(rendered, [["Chosen", "selected-value"]])


if __name__ == "__main__":
    unittest.main()
