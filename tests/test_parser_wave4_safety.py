from __future__ import annotations

from contextlib import contextmanager, ExitStack
from dataclasses import replace
import hashlib
import io
import json
import os
from pathlib import Path
import socket
import tempfile
import unittest
from unittest import mock
import zipfile

from openpyxl import Workbook

from editor_contracts import EditorProject, EditorSegment
from editor_project import ProjectError, save_project
from parser_composition import (
    OpenedParserInput,
    ParserApplicationSurface,
    create_parser_application_surface,
)
from parser_contracts import (
    EffectivePurpose,
    FormatId,
    GETTEXT_PO_V1,
    GETTEXT_POT_V1,
    LINE_TEXT_V1,
    LOCALCAT_JSON_V1,
    NORMALIZED_TM_JSON_V1,
    ParsedSegment,
    ReadRequest,
    ResourceRecord,
    SelectionFailure,
    SelectionRequest,
    SourceReference,
    TERMBASE_CSV_V1,
    TERMBASE_XLSX_V1,
    TMX_LEVEL1_V1,
    TermbaseColumnSelection,
    TermbaseReadOptions,
    TmxReadOptions,
    ValidationOutcome,
)
from parser_gettext_codec import (
    GETTEXT_PO_DESCRIPTOR,
    GETTEXT_POT_DESCRIPTOR,
    GettextPoCodec,
    GettextPotCodec,
)
from parser_localcat_codec import (
    LINE_TEXT_DESCRIPTOR,
    LOCALCAT_JSON_DESCRIPTOR,
    LineTextReader,
    LocalCatJsonReader,
)
from parser_source import (
    CancellationToken,
    GuardedParseSession,
    ParserSessionError,
    ParserSourceError,
    create_sealed_snapshot,
    materialize,
    validate,
)
from parser_termbase_codec import (
    TERMBASE_CSV_DESCRIPTOR,
    TERMBASE_XLSX_DESCRIPTOR,
    CsvTermbaseCodec,
    XlsxTermbaseCodec,
)
from parser_tm_json_codec import (
    NORMALIZED_TM_JSON_DESCRIPTOR,
    NormalizedTmJsonReader,
)
from parser_tmx_codec import TMX_CODEC_DESCRIPTOR, TmxLevel1Codec
from resource_importer import import_termbase, import_tmx
from tests.test_tm_legacy_facade_import import _activate_resource, _store_for
from tm_sqlite_store import SQLiteStoreLifecycleError, SQLiteTMStore


# Task 5.1 deliberately reuses the deeper state-machine and batch fixtures rather
# than cloning them here.  These qualified tests are the adjacent evidence matrix
# run with this Application-facing safety suite: fatal tail / consumer abort and
# cancellation / no verified EOF / terminal exactly once / per-file batch
# independence.  The tests below add the missing eight-format and real commit
# integration coverage.
_TASK_5_1_ADJACENT_EVIDENCE = (
    "tests.test_parser_gettext_codec."
    "GettextFailureAndTerminalTests."
    "test_fatal_tail_can_expose_provisional_record_but_never_terminal",
    "tests.test_parser_tmx_codec."
    "TmxTerminalAndCancellationTests."
    "test_fatal_tail_exposes_provisional_record_but_never_terminal",
    "tests.test_parser_source."
    "GuardedSessionTests."
    "test_early_close_consumer_exception_and_cancellation_abort_session",
    "tests.test_parser_source."
    "GuardedSessionTests."
    "test_unknown_terminal_fatal_tail_unconsumed_and_raw_exception_never_sign",
    "tests.test_parser_source."
    "GuardedSessionTests."
    "test_valid_project_stream_signs_terminal_once_after_true_eof",
    "tests.test_parser_cli_runner_characterization."
    "MigratedNormalizedTMJSONContractTests."
    "test_bad_file_contributes_nothing_while_a_terminal_success_file_can_continue",
)

_EXPECTED_BUILTIN_BINDINGS = frozenset(
    {
        (EffectivePurpose.PROJECT_DOCUMENT, LOCALCAT_JSON_V1),
        (EffectivePurpose.PROJECT_DOCUMENT, LINE_TEXT_V1),
        (EffectivePurpose.PROJECT_DOCUMENT, GETTEXT_PO_V1),
        (EffectivePurpose.PROJECT_DOCUMENT, GETTEXT_POT_V1),
        (EffectivePurpose.TRANSLATION_MEMORY, TMX_LEVEL1_V1),
        (EffectivePurpose.TRANSLATION_MEMORY, NORMALIZED_TM_JSON_V1),
        (EffectivePurpose.TERMBASE, TERMBASE_CSV_V1),
        (EffectivePurpose.TERMBASE, TERMBASE_XLSX_V1),
    }
)

_EXPECTED_FOUNDATION_ISSUES = (
    "PARSER.LIMIT.FIELD",
    "PARSER.LIMIT.INPUT",
    "PARSER.LIMIT.MATERIALIZATION",
    "PARSER.LIMIT.METADATA",
    "PARSER.LIMIT.RECORD",
    "PARSER.PLUGIN.ISSUE_UNDECLARED",
    "PARSER.SOURCE.CANCELLED",
    "PARSER.SOURCE.READ_FAILED",
    "PARSER.SOURCE.STALE",
    "PARSER.SYNTAX.DUPLICATE_LOCAL_ID",
    "PARSER.SYNTAX.INVALID_EVENT",
    "PARSER.SYNTAX.INVALID_HEADER",
    "PARSER.SYNTAX.MALFORMED",
)


def _expected_issues(*additional: str) -> tuple[str, ...]:
    return tuple(sorted((*_EXPECTED_FOUNDATION_ISSUES, *additional)))


def _limit_profile_fact(profile) -> tuple[object, ...]:
    return (
        profile.profile_id,
        profile.profile_version,
        profile.max_input_bytes,
        profile.max_decoded_field_chars,
        profile.max_records,
        profile.max_materialized_records,
        profile.max_retained_issues,
        profile.declared_issue_codes,
        profile.max_metadata_entries_per_container,
        profile.max_metadata_decoded_chars_per_container,
        profile.max_metadata_decoded_chars_total,
        profile.max_structure_depth,
        profile.max_expanded_bytes,
        profile.max_archive_members,
        profile.max_compression_ratio,
    )


_EXPECTED_LIMIT_PROFILE_FACTS = {
    LOCALCAT_JSON_V1: (
        "localcat-json-v1",
        1,
        100 * 1024 * 1024,
        100 * 1024 * 1024,
        100_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.LIMIT.DEPTH",
            "PARSER.SOURCE.ENCODING_FAILED",
            "PARSER.SYNTAX.EMPTY_INPUT",
            "PARSER.SYNTAX.INVALID_FIELD",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        64,
        None,
        None,
        None,
    ),
    LINE_TEXT_V1: (
        "line-text-v1",
        1,
        100 * 1024 * 1024,
        100 * 1024 * 1024,
        1_000_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.SOURCE.ENCODING_FAILED",
            "PARSER.SYNTAX.EMPTY_INPUT",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        8,
        None,
        None,
        None,
    ),
    GETTEXT_PO_V1: (
        "gettext-po-v1",
        1,
        100 * 1024 * 1024,
        100 * 1024 * 1024,
        1_000_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.GETTEXT.CHARSET_UNSUPPORTED",
            "PARSER.GETTEXT.EMPTY_INPUT",
            "PARSER.GETTEXT.PLURAL_UNSUPPORTED",
            "PARSER.GETTEXT.SYNTAX",
            "PARSER.SOURCE.ENCODING_FAILED",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        16,
        None,
        None,
        None,
    ),
    GETTEXT_POT_V1: (
        "gettext-pot-v1",
        1,
        100 * 1024 * 1024,
        100 * 1024 * 1024,
        1_000_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.GETTEXT.CHARSET_UNSUPPORTED",
            "PARSER.GETTEXT.EMPTY_INPUT",
            "PARSER.GETTEXT.PLURAL_UNSUPPORTED",
            "PARSER.GETTEXT.SYNTAX",
            "PARSER.SOURCE.ENCODING_FAILED",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        16,
        None,
        None,
        None,
    ),
    TMX_LEVEL1_V1: (
        "tmx-level1-v1",
        1,
        100 * 1024 * 1024,
        1_000_000,
        1_000_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.LIMIT.DEPTH",
            "PARSER.SOURCE.ENCODING_FAILED",
            "PARSER.TMX.INLINE_XML_UNSUPPORTED",
            "PARSER.TMX.LOCALE_FALLBACK_AMBIGUOUS",
            "PARSER.TMX.LOCALE_PAIR_MISSING",
            "PARSER.TMX.LOCALE_SELECTION_INVALID",
            "PARSER.TMX.NO_TRANSLATION_UNITS",
            "PARSER.TMX.SEGMENT_LIMIT",
            "PARSER.TMX.UNSAFE_XML",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        64,
        None,
        None,
        None,
    ),
    NORMALIZED_TM_JSON_V1: (
        "normalized-tm-json-v1",
        1,
        100 * 1024 * 1024,
        100 * 1024 * 1024,
        100_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.LIMIT.DEPTH",
            "PARSER.SOURCE.ENCODING_FAILED",
            "PARSER.SYNTAX.EMPTY_INPUT",
            "PARSER.SYNTAX.INVALID_FIELD",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        64,
        None,
        None,
        None,
    ),
    TERMBASE_CSV_V1: (
        "termbase-csv-v1",
        1,
        100 * 1024 * 1024,
        100 * 1024 * 1024,
        1_000_000,
        100_000,
        256,
        _expected_issues(
            "PARSER.CAPABILITY.CONDITIONAL_DEPENDENCY_INCOMPATIBLE",
            "PARSER.CAPABILITY.CONDITIONAL_DEPENDENCY_MISSING",
            "PARSER.SOURCE.ENCODING_FAILED",
            "PARSER.TERMBASE.ACTIVE_SHEET_MISSING",
            "PARSER.TERMBASE.COLUMN_SELECTION_SAME",
            "PARSER.TERMBASE.HEADER_DUPLICATE",
            "PARSER.TERMBASE.HEADER_MISSING",
            "PARSER.TERMBASE.HEADER_SKIPPED",
            "PARSER.TERMBASE.ROW_EMPTY",
            "PARSER.TERMBASE.ROW_MISSING_COLUMN",
            "PARSER.TERMBASE.SOURCE_EMPTY",
            "PARSER.TERMBASE.TARGET_EMPTY",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        8,
        None,
        None,
        None,
    ),
    TERMBASE_XLSX_V1: (
        "termbase-xlsx-v1",
        1,
        100 * 1024 * 1024,
        1_000_000,
        1_048_576,
        100_000,
        256,
        _expected_issues(
            "PARSER.CAPABILITY.CONDITIONAL_DEPENDENCY_INCOMPATIBLE",
            "PARSER.CAPABILITY.CONDITIONAL_DEPENDENCY_MISSING",
            "PARSER.LIMIT.ARCHIVE_MEMBER",
            "PARSER.LIMIT.COMPRESSION_RATIO",
            "PARSER.LIMIT.EXPANSION",
            "PARSER.LIMIT.STRUCTURE_DEPTH",
            "PARSER.SOURCE.ENCODING_FAILED",
            "PARSER.TERMBASE.ACTIVE_SHEET_MISSING",
            "PARSER.TERMBASE.COLUMN_SELECTION_SAME",
            "PARSER.TERMBASE.HEADER_DUPLICATE",
            "PARSER.TERMBASE.HEADER_MISSING",
            "PARSER.TERMBASE.HEADER_SKIPPED",
            "PARSER.TERMBASE.ROW_EMPTY",
            "PARSER.TERMBASE.ROW_MISSING_COLUMN",
            "PARSER.TERMBASE.SOURCE_EMPTY",
            "PARSER.TERMBASE.TARGET_EMPTY",
            "PARSER.TERMBASE.XML_DECLARATION_FORBIDDEN",
            "PARSER.XLSX.ARCHIVE_DATA_DESCRIPTOR_UNSUPPORTED",
            "PARSER.XLSX.ARCHIVE_INVALID",
            "PARSER.XLSX.ARCHIVE_MEMBER_DUPLICATE",
            "PARSER.XLSX.ARCHIVE_MEMBER_NAME_UNSAFE",
            "PARSER.XLSX.SOURCE_NOT_SEEKABLE",
            "PARSER.XLSX.SOURCE_RESTORE_FAILED",
        ),
        256,
        1024 * 1024,
        16 * 1024 * 1024,
        64,
        256 * 1024 * 1024,
        4096,
        100.0,
    ),
}


@contextmanager
def _record_real_resource_parser_bridge():
    original_open = ParserApplicationSurface.open_input
    original_stream = OpenedParserInput.stream
    original_terminal = GuardedParseSession.verified_terminal
    with mock.patch(
        "resource_importer.create_parser_application_surface",
        wraps=create_parser_application_surface,
    ) as factory, mock.patch.object(
        ParserApplicationSurface,
        "open_input",
        autospec=True,
        side_effect=original_open,
    ) as open_input, mock.patch.object(
        OpenedParserInput,
        "stream",
        autospec=True,
        side_effect=original_stream,
    ) as stream, mock.patch.object(
        GuardedParseSession,
        "verified_terminal",
        autospec=True,
        side_effect=original_terminal,
    ) as terminal:
        yield factory, open_input, stream, terminal


class _InputCase:
    def __init__(
        self,
        filename: str,
        format_id: FormatId,
        purpose: EffectivePurpose,
        payload: bytes,
    ) -> None:
        self.filename = filename
        self.format_id = format_id
        self.purpose = purpose
        self.payload = payload

    def request(self) -> ReadRequest:
        if self.format_id == TMX_LEVEL1_V1:
            return ReadRequest(
                purpose=self.purpose,
                format_id=self.format_id,
                tmx_options=TmxReadOptions("en-US", "zh-CN"),
            )
        if self.format_id in (TERMBASE_CSV_V1, TERMBASE_XLSX_V1):
            return ReadRequest(
                purpose=self.purpose,
                format_id=self.format_id,
                termbase_options=TermbaseReadOptions(
                    TermbaseColumnSelection.legacy_first_two_columns()
                ),
            )
        return ReadRequest(purpose=self.purpose, format_id=self.format_id)


def _xlsx_bytes(rows: tuple[tuple[object, object], ...]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _valid_cases() -> tuple[_InputCase, ...]:
    return (
        _InputCase(
            "project.json",
            LOCALCAT_JSON_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'[{"source":"A"}]',
        ),
        _InputCase(
            "project.txt",
            LINE_TEXT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b"A\n",
        ),
        _InputCase(
            "project.po",
            GETTEXT_PO_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'msgid "A"\nmsgstr "B"\n',
        ),
        _InputCase(
            "project.pot",
            GETTEXT_POT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'msgid "A"\nmsgstr ""\n',
        ),
        _InputCase(
            "memory.tmx",
            TMX_LEVEL1_V1,
            EffectivePurpose.TRANSLATION_MEMORY,
            (
                b'<tmx><body><tu><tuv xml:lang="en-US"><seg>A</seg></tuv>'
                b'<tuv xml:lang="zh-CN"><seg>B</seg></tuv></tu></body></tmx>'
            ),
        ),
        _InputCase(
            "memory.json",
            NORMALIZED_TM_JSON_V1,
            EffectivePurpose.TRANSLATION_MEMORY,
            b'[{"source":"A","target":"B"}]',
        ),
        _InputCase(
            "terms.csv",
            TERMBASE_CSV_V1,
            EffectivePurpose.TERMBASE,
            b"Source,Target\nA,B\n",
        ),
        _InputCase(
            "terms.xlsx",
            TERMBASE_XLSX_V1,
            EffectivePurpose.TERMBASE,
            _xlsx_bytes((("Source", "Target"), ("A", "B"))),
        ),
    )


def _invalid_encoding_cases() -> tuple[_InputCase, ...]:
    invalid_xlsx = io.BytesIO()
    with zipfile.ZipFile(
        invalid_xlsx,
        "w",
        compression=zipfile.ZIP_STORED,
    ) as archive:
        archive.writestr("[Content_Types].xml", b"<Types>\xff</Types>")
        archive.writestr("xl/workbook.xml", b"<workbook>\xff</workbook>")
    return (
        _InputCase(
            "invalid.json",
            LOCALCAT_JSON_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'[{"source":"\xff"}]',
        ),
        _InputCase(
            "invalid.txt",
            LINE_TEXT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b"\xff\n",
        ),
        _InputCase(
            "invalid.po",
            GETTEXT_PO_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'msgid "\xff"\nmsgstr "B"\n',
        ),
        _InputCase(
            "invalid.pot",
            GETTEXT_POT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'msgid "\xff"\nmsgstr ""\n',
        ),
        _InputCase(
            "invalid.tmx",
            TMX_LEVEL1_V1,
            EffectivePurpose.TRANSLATION_MEMORY,
            b'<tmx><body><tu><tuv xml:lang="en-US"><seg>\xff</seg></tuv>'
            b'<tuv xml:lang="zh-CN"><seg>B</seg></tuv></tu></body></tmx>',
        ),
        _InputCase(
            "invalid-memory.json",
            NORMALIZED_TM_JSON_V1,
            EffectivePurpose.TRANSLATION_MEMORY,
            b'[{"source":"\xff","target":"B"}]',
        ),
        _InputCase(
            "invalid.csv",
            TERMBASE_CSV_V1,
            EffectivePurpose.TERMBASE,
            b"\xff,B\n",
        ),
        _InputCase(
            "invalid.xlsx",
            TERMBASE_XLSX_V1,
            EffectivePurpose.TERMBASE,
            invalid_xlsx.getvalue(),
        ),
    )


def _two_record_cases() -> tuple[_InputCase, ...]:
    return (
        _InputCase(
            "two.json",
            LOCALCAT_JSON_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'[{"source":"Long"},{"source":"Other"}]',
        ),
        _InputCase(
            "two.txt",
            LINE_TEXT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b"Long\nOther\n",
        ),
        _InputCase(
            "two.po",
            GETTEXT_PO_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'msgid "Long"\nmsgstr "B"\n\nmsgid "Other"\nmsgstr "D"\n',
        ),
        _InputCase(
            "two.pot",
            GETTEXT_POT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b'msgid "Long"\nmsgstr ""\n\nmsgid "Other"\nmsgstr ""\n',
        ),
        _InputCase(
            "two.tmx",
            TMX_LEVEL1_V1,
            EffectivePurpose.TRANSLATION_MEMORY,
            b'<tmx><body><tu><tuv xml:lang="en-US"><seg>Long</seg></tuv>'
            b'<tuv xml:lang="zh-CN"><seg>B</seg></tuv></tu>'
            b'<tu><tuv xml:lang="en-US"><seg>Other</seg></tuv>'
            b'<tuv xml:lang="zh-CN"><seg>D</seg></tuv></tu></body></tmx>',
        ),
        _InputCase(
            "two-memory.json",
            NORMALIZED_TM_JSON_V1,
            EffectivePurpose.TRANSLATION_MEMORY,
            b'[{"source":"Long","target":"B"},'
            b'{"source":"Other","target":"D"}]',
        ),
        _InputCase(
            "two.csv",
            TERMBASE_CSV_V1,
            EffectivePurpose.TERMBASE,
            b"Long,B\nOther,D\n",
        ),
        _InputCase(
            "two.xlsx",
            TERMBASE_XLSX_V1,
            EffectivePurpose.TERMBASE,
            _xlsx_bytes((("Long", "B"), ("Other", "D"))),
        ),
    )


def _xlsx_with_inert_features() -> bytes:
    """Return a valid workbook carrying features the reader must never execute."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("Source", "Target", "Formula"))
    worksheet.append(
        (
            "A",
            "B",
            '=WEBSERVICE("https://invalid.example/formula")',
        )
    )
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    source = zipfile.ZipFile(io.BytesIO(output.getvalue()), "r")
    enriched = io.BytesIO()
    with source, zipfile.ZipFile(
        enriched,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as target:
        contents = {
            member.filename: source.read(member.filename)
            for member in source.infolist()
        }
        content_types = contents.pop("[Content_Types].xml")
        content_types = content_types.replace(
            b"</Types>",
            (
                b'<Override PartName="/xl/vbaProject.bin" '
                b'ContentType="application/vnd.ms-office.vbaProject"/>'
                b'<Override PartName="/xl/embeddings/oleObject1.bin" '
                b'ContentType="application/vnd.openxmlformats-officedocument.'
                b'oleObject"/>'
                b'<Override PartName="/xl/externalLinks/externalLink1.xml" '
                b'ContentType="application/vnd.openxmlformats-officedocument.'
                b'spreadsheetml.externalLink+xml"/></Types>'
            ),
        )
        contents["[Content_Types].xml"] = content_types

        workbook_xml = contents.pop("xl/workbook.xml")
        if b"xmlns:r=" not in workbook_xml:
            workbook_xml = workbook_xml.replace(
                b"<workbook ",
                b'<workbook xmlns:r="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships" ',
                1,
            )
        workbook_xml = workbook_xml.replace(
            b"</workbook>",
            b'<externalReferences><externalReference r:id="rIdExternal"/>'
            b"</externalReferences></workbook>",
        )
        contents["xl/workbook.xml"] = workbook_xml

        workbook_rels = contents.pop("xl/_rels/workbook.xml.rels")
        workbook_rels = workbook_rels.replace(
            b"</Relationships>",
            (
                b'<Relationship Id="rIdExternal" '
                b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships/externalLink" '
                b'Target="externalLinks/externalLink1.xml"/>'
                b'<Relationship Id="rIdVba" '
                b'Type="http://schemas.microsoft.com/office/2006/relationships/'
                b'vbaProject" Target="vbaProject.bin"/></Relationships>'
            ),
        )
        contents["xl/_rels/workbook.xml.rels"] = workbook_rels

        worksheet_xml = contents.pop("xl/worksheets/sheet1.xml")
        if b"xmlns:r=" not in worksheet_xml:
            worksheet_xml = worksheet_xml.replace(
                b"<worksheet ",
                b'<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships" ',
                1,
            )
        worksheet_xml = worksheet_xml.replace(
            b"</worksheet>",
            b'<oleObjects><oleObject progId="Package" shapeId="1" '
            b'r:id="rIdOle"/></oleObjects></worksheet>',
        )
        contents["xl/worksheets/sheet1.xml"] = worksheet_xml

        for name, payload in contents.items():
            target.writestr(name, payload)
        target.writestr("xl/vbaProject.bin", b"INERT-MACRO")
        target.writestr("xl/embeddings/oleObject1.bin", b"INERT-OBJECT")
        target.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/'
            b'2006/relationships"><Relationship Id="rIdOle" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships/oleObject" '
            b'Target="../embeddings/oleObject1.bin"/></Relationships>',
        )
        target.writestr(
            "xl/externalLinks/externalLink1.xml",
            b'<externalLink xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main" '
            b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships"><externalBook r:id="rId1"/></externalLink>',
        )
        target.writestr(
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/'
            b'2006/relationships"><Relationship Id="rId1" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships/externalLinkPath" Target="https://invalid.example/'
            b'external.xlsx" TargetMode="External"/></Relationships>',
        )
    return enriched.getvalue()


class ParserWave4SafetyTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary.cleanup)
        self.root = Path(self.temporary.name)
        self.surface = create_parser_application_surface()

    def _write(self, case: _InputCase) -> Path:
        path = self.root / case.filename
        path.write_bytes(case.payload)
        return path

    @staticmethod
    def _tracking_snapshot_files():
        actual_temporary_file = tempfile.TemporaryFile
        handles = []

        def tracking_temporary_file(*args, **kwargs):
            handle = actual_temporary_file(*args, **kwargs)
            handles.append(handle)
            return handle

        return (
            mock.patch(
                "parser_source.tempfile.TemporaryFile",
                side_effect=tracking_temporary_file,
            ),
            handles,
        )

    def _assert_snapshot_handles_closed(
        self,
        handles,
        *,
        expected_count: int,
    ) -> None:
        self.assertEqual(len(handles), expected_count)
        self.assertEqual([handle for handle in handles if not handle.closed], [])
        self.assertTrue(all(handle.closed for handle in handles))

    def _assert_resource_parser_bridge(
        self,
        bridge,
        *,
        purpose: EffectivePurpose,
        format_id: FormatId,
        expected_count: int = 1,
    ) -> None:
        factory, open_input, stream, terminal = bridge
        self.assertEqual(factory.call_count, expected_count)
        self.assertTrue(all(call.args == () and call.kwargs == {} for call in factory.call_args_list))
        self.assertEqual(open_input.call_count, expected_count)
        self.assertEqual(stream.call_count, expected_count)
        self.assertEqual(terminal.call_count, expected_count)
        for call in open_input.call_args_list:
            open_args, open_kwargs = call
            self.assertEqual(open_kwargs, {})
            self.assertEqual(len(open_args), 4)
            self.assertIs(type(open_args[0]), ParserApplicationSurface)
            self.assertIs(type(open_args[1]), SourceReference)
            self.assertIs(type(open_args[2]), SelectionRequest)
            self.assertIs(type(open_args[3]), ReadRequest)
            self.assertIs(open_args[2].purpose, purpose)
            self.assertEqual(open_args[2].format_id, format_id)
            self.assertIs(open_args[3].purpose, purpose)
            self.assertEqual(open_args[3].format_id, format_id)
        self.assertTrue(
            all(type(call.args[0]) is OpenedParserInput for call in stream.call_args_list)
        )
        self.assertTrue(
            all(
                type(call.args[0]) is GuardedParseSession
                for call in terminal.call_args_list
            )
        )

    def _assert_complete_builtin_matrix(
        self,
        cases: tuple[_InputCase, ...],
    ) -> None:
        bindings = tuple((case.purpose, case.format_id) for case in cases)
        self.assertEqual(len(bindings), 8)
        self.assertEqual(len(set(bindings)), len(bindings))
        self.assertEqual(frozenset(bindings), _EXPECTED_BUILTIN_BINDINGS)

    def test_input_fixture_matrices_are_independent_complete_builtin_bindings(
        self,
    ) -> None:
        matrices = {
            "valid": _valid_cases(),
            "invalid-encoding": _invalid_encoding_cases(),
            "two-record": _two_record_cases(),
        }
        self.assertEqual(set(matrices), {"valid", "invalid-encoding", "two-record"})
        self.assertEqual(len({id(cases) for cases in matrices.values()}), 3)
        for label, cases in matrices.items():
            with self.subTest(matrix=label):
                self._assert_complete_builtin_matrix(cases)
                self.assertEqual(
                    len({case.filename for case in cases}),
                    len(cases),
                )

    def test_task_5_1_adjacent_evidence_matrix_executes(self) -> None:
        suite = unittest.defaultTestLoader.loadTestsFromNames(
            _TASK_5_1_ADJACENT_EVIDENCE
        )
        result = unittest.TestResult()
        suite.run(result)
        self.assertEqual(result.testsRun, len(_TASK_5_1_ADJACENT_EVIDENCE))
        self.assertEqual(len(result.failures), 0, result.failures)
        self.assertEqual(len(result.errors), 0, result.errors)

    def _open(
        self,
        case: _InputCase,
        path: Path | None = None,
        *,
        cancellation: CancellationToken | None = None,
    ):
        selected = self._write(case) if path is None else path
        opened = self.surface.open_input(
            SourceReference(
                safe_root=str(self.root),
                selected_path=str(selected),
                display_hint=case.filename,
            ),
            SelectionRequest(case.purpose, format_id=case.format_id),
            case.request(),
            cancellation=cancellation,
        )
        self.assertNotIsInstance(opened, SelectionFailure)
        return opened

    def _direct_reader(self, case, descriptor, stack: ExitStack):
        readers = {
            LOCALCAT_JSON_V1: LocalCatJsonReader,
            LINE_TEXT_V1: LineTextReader,
            GETTEXT_PO_V1: GettextPoCodec,
            GETTEXT_POT_V1: GettextPotCodec,
            NORMALIZED_TM_JSON_V1: NormalizedTmJsonReader,
            TERMBASE_CSV_V1: CsvTermbaseCodec,
        }
        if case.format_id == TMX_LEVEL1_V1:
            stack.enter_context(
                mock.patch("parser_tmx_codec.TMX_CODEC_DESCRIPTOR", descriptor)
            )
            return TmxLevel1Codec()
        if case.format_id == TERMBASE_XLSX_V1:
            stack.enter_context(
                mock.patch(
                    "parser_termbase_codec.TERMBASE_XLSX_DESCRIPTOR",
                    descriptor,
                )
            )
            return XlsxTermbaseCodec()
        return readers[case.format_id](descriptor)

    def _direct_snapshot(self, case, descriptor):
        selected = self._write(case)
        return create_sealed_snapshot(
            SourceReference(
                safe_root=str(self.root),
                selected_path=str(selected),
                display_hint=case.filename,
            ),
            limit_profile=descriptor.limit_profile,
        )

    def _assert_inert_xlsx_package(self, payload: bytes) -> None:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
            names = set(archive.namelist())
            required = {
                "[Content_Types].xml",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
                "xl/worksheets/sheet1.xml",
                "xl/worksheets/_rels/sheet1.xml.rels",
                "xl/vbaProject.bin",
                "xl/embeddings/oleObject1.bin",
                "xl/externalLinks/externalLink1.xml",
                "xl/externalLinks/_rels/externalLink1.xml.rels",
            }
            self.assertTrue(required <= names)
            content_types = archive.read("[Content_Types].xml")
            workbook = archive.read("xl/workbook.xml")
            workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
            worksheet = archive.read("xl/worksheets/sheet1.xml")
            worksheet_rels = archive.read(
                "xl/worksheets/_rels/sheet1.xml.rels"
            )
            external_link = archive.read("xl/externalLinks/externalLink1.xml")
            external_rels = archive.read(
                "xl/externalLinks/_rels/externalLink1.xml.rels"
            )

        self.assertIn(b"<f>WEBSERVICE", worksheet)
        self.assertIn(b"invalid.example/formula", worksheet)
        self.assertIn(b"<oleObjects>", worksheet)
        self.assertIn(b'r:id="rIdOle"', worksheet)
        self.assertIn(b"<externalReferences>", workbook)
        self.assertIn(b'r:id="rIdExternal"', workbook)
        for part_name in (
            b'/xl/vbaProject.bin',
            b'/xl/embeddings/oleObject1.bin',
            b'/xl/externalLinks/externalLink1.xml',
        ):
            self.assertIn(part_name, content_types)
        self.assertIn(b'Target="vbaProject.bin"', workbook_rels)
        self.assertIn(
            b'Target="externalLinks/externalLink1.xml"',
            workbook_rels,
        )
        self.assertIn(
            b'Target="../embeddings/oleObject1.bin"',
            worksheet_rels,
        )
        self.assertIn(b'<externalBook r:id="rId1"', external_link)
        self.assertIn(b'TargetMode="External"', external_rels)
        self.assertIn(b"invalid.example/external.xlsx", external_rels)

    def test_all_builtin_views_bind_terminal_to_the_actual_profile_and_snapshot(
        self,
    ) -> None:
        self._assert_complete_builtin_matrix(_valid_cases())
        expected_record_limits = {
            LOCALCAT_JSON_V1: 100_000,
            LINE_TEXT_V1: 1_000_000,
            GETTEXT_PO_V1: 1_000_000,
            GETTEXT_POT_V1: 1_000_000,
            TMX_LEVEL1_V1: 1_000_000,
            NORMALIZED_TM_JSON_V1: 100_000,
            TERMBASE_CSV_V1: 1_000_000,
            TERMBASE_XLSX_V1: 1_048_576,
        }
        observed_limits: set[int] = set()

        for case in _valid_cases():
            with self.subTest(format_id=case.format_id.value):
                with self._open(case) as opened:
                    validation = opened.validate()
                    materialized = opened.materialize()

                    self.assertIs(validation.outcome, ValidationOutcome.SUCCESS)
                    self.assertIsNotNone(validation.terminal)
                    self.assertEqual(validation.source, opened.source_identity)
                    self.assertEqual(
                        materialized.terminal.source,
                        opened.source_identity,
                    )
                    self.assertEqual(
                        validation.codec_identity,
                        opened.descriptor.identity,
                    )
                    self.assertEqual(
                        validation.terminal.codec_identity,
                        opened.descriptor.identity,
                    )
                    self.assertEqual(
                        validation.limit_profile,
                        opened.descriptor.limit_profile,
                    )
                    self.assertEqual(
                        materialized.terminal.limit_profile,
                        opened.descriptor.limit_profile,
                    )
                    self.assertEqual(
                        validation.limit_profile.max_records,
                        expected_record_limits[case.format_id],
                    )
                    self.assertEqual(
                        validation.limit_profile.max_materialized_records,
                        100_000,
                    )
                    self.assertEqual(
                        validation.terminal.record_count,
                        materialized.terminal.record_count,
                    )
                    observed_limits.add(validation.limit_profile.max_records)

        # Gate D's 100k TM query qualification is not a universal Parser
        # record limit: Parser publishes independent, format-owned profiles.
        self.assertEqual(observed_limits, {100_000, 1_000_000, 1_048_576})

    def test_all_builtin_streams_destroy_provisional_state_on_early_close(
        self,
    ) -> None:
        self._assert_complete_builtin_matrix(_valid_cases())
        all_handles = []
        for case in _valid_cases():
            with self.subTest(format_id=case.format_id.value):
                temporary_patch, handles = self._tracking_snapshot_files()
                with temporary_patch:
                    with self._open(case) as opened:
                        snapshot = opened._snapshot
                        session = opened.stream()
                        iterator = iter(session)
                        first_record = None
                        for event in iterator:
                            if type(event) in {ParsedSegment, ResourceRecord}:
                                first_record = event
                                break

                        self.assertIsNotNone(first_record)
                        self.assertEqual(session.provisional_record_count, 1)
                        iterator.close()

                        with self.assertRaises(ParserSessionError) as caught:
                            session.verified_terminal()
                        self.assertEqual(
                            caught.exception.code,
                            "PARSER.SESSION.ABORTED",
                        )
                        session.close()
                        self.assertTrue(session.source.closed)
                self.assertTrue(snapshot.released)
                self._assert_snapshot_handles_closed(handles, expected_count=1)
                all_handles.extend(handles)
                self.assertEqual(tuple(self.root.glob("parser-snapshot-*")), ())
        self._assert_snapshot_handles_closed(all_handles, expected_count=8)

    def test_all_builtin_cancellation_paths_deny_terminal_and_release_snapshot(
        self,
    ) -> None:
        self._assert_complete_builtin_matrix(_valid_cases())
        all_handles = []
        for case in _valid_cases():
            with self.subTest(format_id=case.format_id.value):
                temporary_patch, handles = self._tracking_snapshot_files()
                with temporary_patch:
                    cancellation = CancellationToken()
                    opened = self._open(case, cancellation=cancellation)
                    snapshot = opened._snapshot
                    session = opened.stream()
                    iterator = iter(session)
                    first_record = None
                    for event in iterator:
                        if type(event) in {ParsedSegment, ResourceRecord}:
                            first_record = event
                            break
                    self.assertIsNotNone(first_record)
                    self.assertEqual(session.provisional_record_count, 1)

                    cancellation.cancel()
                    with self.assertRaises(ParserSessionError) as caught:
                        next(iterator)
                    self.assertEqual(
                        caught.exception.code,
                        "PARSER.SOURCE.CANCELLED",
                    )
                    with self.assertRaises(ParserSessionError):
                        session.verified_terminal()

                    iterator.close()
                    session.close()
                    opened.close()
                    self.assertTrue(session.source.closed)
                    self.assertTrue(snapshot.released)
                    with self.assertRaises(ParserSourceError) as released:
                        opened.stream()
                    self.assertEqual(
                        released.exception.code,
                        "PARSER.SOURCE.SNAPSHOT_RELEASED",
                    )
                self._assert_snapshot_handles_closed(handles, expected_count=1)
                all_handles.extend(handles)
        self._assert_snapshot_handles_closed(all_handles, expected_count=8)

    def test_all_builtin_normal_views_close_the_real_snapshot_handles(self) -> None:
        cases = _valid_cases()
        self._assert_complete_builtin_matrix(cases)
        temporary_patch, handles = self._tracking_snapshot_files()
        with temporary_patch:
            for case in cases:
                with self.subTest(format_id=case.format_id.value):
                    with self._open(case) as opened:
                        result = opened.materialize()
                        self.assertEqual(result.terminal.record_count, 1)
        self._assert_snapshot_handles_closed(handles, expected_count=8)

    @unittest.skipUnless(os.name == "posix", "rooted no-follow matrix is POSIX-specific")
    def test_all_builtin_entries_reject_final_symlinks_before_a_guarded_view(
        self,
    ) -> None:
        self._assert_complete_builtin_matrix(_valid_cases())
        for index, case in enumerate(_valid_cases()):
            with self.subTest(format_id=case.format_id.value):
                real = self._write(case)
                link = self.root / f"link-{index}{real.suffix}"
                link.symlink_to(real)

                with mock.patch(
                    "parser_source.tempfile.TemporaryFile",
                    wraps=tempfile.TemporaryFile,
                ) as temporary_file:
                    with self.assertRaises(ParserSourceError) as caught:
                        self._open(case, link)
                self.assertEqual(caught.exception.code, "PARSER.SOURCE.NOT_REGULAR")
                temporary_file.assert_not_called()

    def test_live_snapshot_survives_path_mutation_and_a_new_open_sees_new_bytes(
        self,
    ) -> None:
        case = next(item for item in _valid_cases() if item.format_id == LINE_TEXT_V1)
        path = self._write(case)

        with self._open(case, path) as opened:
            original_digest = opened.source_identity.content_sha256
            validation = opened.validate()
            path.write_bytes(b"changed\n")
            materialized = opened.materialize()

            self.assertIs(validation.outcome, ValidationOutcome.SUCCESS)
            self.assertEqual(materialized.records[0].source, "A")
            self.assertEqual(materialized.terminal.source, validation.source)
            self.assertEqual(
                materialized.terminal.source.content_sha256,
                original_digest,
            )

        with self._open(case, path) as reopened:
            replacement = reopened.materialize()
            self.assertEqual(replacement.records[0].source, "changed")
            self.assertEqual(
                replacement.terminal.source.content_sha256,
                hashlib.sha256(b"changed\n").hexdigest(),
            )
            self.assertNotEqual(
                replacement.terminal.source.content_sha256,
                original_digest,
            )

    @unittest.skipUnless(os.name == "posix", "rooted no-follow matrix is POSIX-specific")
    def test_application_open_rejects_a_directory_component_swapped_to_symlink(
        self,
    ) -> None:
        selected_dir = self.root / "selected"
        selected_dir.mkdir()
        selected = selected_dir / "chapter.txt"
        selected.write_bytes(b"trusted\n")
        outside = self.root / "outside"
        outside.mkdir()
        (outside / "chapter.txt").write_bytes(b"untrusted\n")
        from parser_source import _open_absolute_root

        swapped = False

        def swap_after_root_anchor(safe_root):
            nonlocal swapped
            root_descriptor = _open_absolute_root(safe_root)
            if not swapped:
                selected_dir.rename(self.root / "selected-original")
                selected_dir.symlink_to(outside, target_is_directory=True)
                swapped = True
            return root_descriptor

        case = _InputCase(
            "chapter.txt",
            LINE_TEXT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b"trusted\n",
        )
        temporary_handles = []
        original_temporary_file = tempfile.TemporaryFile

        def capture_temporary(*args, **kwargs):
            handle = original_temporary_file(*args, **kwargs)
            temporary_handles.append(handle)
            return handle

        with mock.patch(
            "parser_source._open_absolute_root",
            side_effect=swap_after_root_anchor,
        ), mock.patch(
            "parser_source.tempfile.TemporaryFile",
            side_effect=capture_temporary,
        ) as temporary_file:
            with self.assertRaises(ParserSourceError) as caught:
                self._open(case, selected)

        self.assertTrue(swapped)
        self.assertEqual(caught.exception.code, "PARSER.SOURCE.NOT_REGULAR")
        temporary_file.assert_not_called()
        self.assertEqual(temporary_handles, [])
        self.assertTrue(all(handle.closed for handle in temporary_handles))
        self.assertEqual(tuple(self.root.glob("parser-snapshot-*")), ())

    def test_application_open_stale_copy_releases_its_temporary_snapshot(
        self,
    ) -> None:
        case = _InputCase(
            "stale.txt",
            LINE_TEXT_V1,
            EffectivePurpose.PROJECT_DOCUMENT,
            b"A\n",
        )
        selected = self._write(case)
        original_read = os.read
        original_temporary_file = tempfile.TemporaryFile
        temporary_handles = []
        mutated = False

        def capture_temporary(*args, **kwargs):
            handle = original_temporary_file(*args, **kwargs)
            temporary_handles.append(handle)
            return handle

        def mutate_during_copy(descriptor, byte_count):
            nonlocal mutated
            chunk = original_read(descriptor, byte_count)
            if chunk and not mutated:
                selected.write_bytes(b"replacement-with-different-identity\n")
                mutated = True
            return chunk

        with mock.patch(
            "parser_source.tempfile.TemporaryFile",
            side_effect=capture_temporary,
        ), mock.patch("parser_source.os.read", side_effect=mutate_during_copy):
            with self.assertRaises(ParserSourceError) as caught:
                self._open(case, selected)

        self.assertTrue(mutated)
        self.assertEqual(caught.exception.code, "PARSER.SOURCE.STALE")
        self._assert_snapshot_handles_closed(
            temporary_handles,
            expected_count=1,
        )
        self.assertEqual(tuple(self.root.glob("parser-snapshot-*")), ())

    def test_resource_warning_floods_keep_bounded_details_and_exact_counts(
        self,
    ) -> None:
        warning_count = 257
        inline = (
            '<tu><tuv xml:lang="en-US"><seg>LEAK-ME<ph id="1"/></seg></tuv>'
            '<tuv xml:lang="zh-CN"><seg>secret</seg></tuv></tu>'
        )
        cases = (
            (
                _InputCase(
                    "warnings.json",
                    NORMALIZED_TM_JSON_V1,
                    EffectivePurpose.TRANSLATION_MEMORY,
                    json.dumps(
                        [{"source": "A", "target": "B"}]
                        + [{} for _ in range(warning_count)]
                    ).encode("utf-8"),
                ),
                "PARSER.SYNTAX.INVALID_FIELD",
            ),
            (
                _InputCase(
                    "warnings.csv",
                    TERMBASE_CSV_V1,
                    EffectivePurpose.TERMBASE,
                    b"A,B\n" + (b",\n" * warning_count),
                ),
                "PARSER.TERMBASE.ROW_EMPTY",
            ),
            (
                _InputCase(
                    "warnings.tmx",
                    TMX_LEVEL1_V1,
                    EffectivePurpose.TRANSLATION_MEMORY,
                    (
                        '<tmx><body><tu><tuv xml:lang="en-US"><seg>A</seg></tuv>'
                        '<tuv xml:lang="zh-CN"><seg>B</seg></tuv></tu>'
                        + (inline * warning_count)
                        + "</body></tmx>"
                    ).encode("utf-8"),
                ),
                "PARSER.TMX.INLINE_XML_UNSUPPORTED",
            ),
            (
                _InputCase(
                    "warnings.xlsx",
                    TERMBASE_XLSX_V1,
                    EffectivePurpose.TERMBASE,
                    _xlsx_bytes(
                        (("A", "B"),)
                        + tuple(("LEAK-ME", None) for _ in range(warning_count))
                    ),
                ),
                "PARSER.TERMBASE.ROW_MISSING_COLUMN",
            ),
        )

        for case, expected_code in cases:
            with self.subTest(format_id=case.format_id.value):
                with self._open(case) as opened:
                    report = opened.validate()

                counts = {item.code: item.count for item in report.issue_counts}
                self.assertIs(report.outcome, ValidationOutcome.SUCCESS)
                self.assertIsNotNone(report.terminal)
                self.assertTrue(report.issues_truncated)
                self.assertTrue(report.terminal.issues_truncated)
                self.assertEqual(len(report.issues), 256)
                self.assertEqual(counts[expected_code], warning_count)
                self.assertEqual(
                    dict(
                        (item.code, item.count)
                        for item in report.terminal.warning_counts
                    )[expected_code],
                    warning_count,
                )
                self.assertNotIn(
                    "LEAK-ME",
                    " ".join(issue.safe_summary for issue in report.issues),
                )

    def test_all_descriptors_enforce_reduced_common_limit_profiles(
        self,
    ) -> None:
        two_record_cases = _two_record_cases()
        self._assert_complete_builtin_matrix(two_record_cases)
        descriptors = {
            LOCALCAT_JSON_V1: LOCALCAT_JSON_DESCRIPTOR,
            LINE_TEXT_V1: LINE_TEXT_DESCRIPTOR,
            GETTEXT_PO_V1: GETTEXT_PO_DESCRIPTOR,
            GETTEXT_POT_V1: GETTEXT_POT_DESCRIPTOR,
            TMX_LEVEL1_V1: TMX_CODEC_DESCRIPTOR,
            NORMALIZED_TM_JSON_V1: NORMALIZED_TM_JSON_DESCRIPTOR,
            TERMBASE_CSV_V1: TERMBASE_CSV_DESCRIPTOR,
            TERMBASE_XLSX_V1: TERMBASE_XLSX_DESCRIPTOR,
        }
        expected_formats = {format_id for _purpose, format_id in _EXPECTED_BUILTIN_BINDINGS}
        self.assertEqual(set(descriptors), expected_formats)
        input_expectations = {format_id: "PARSER.LIMIT.INPUT" for format_id in expected_formats}
        field_expectations = {
            LOCALCAT_JSON_V1: (ValidationOutcome.FAILED, 0, "PARSER.LIMIT.FIELD"),
            LINE_TEXT_V1: (ValidationOutcome.FAILED, 0, "PARSER.LIMIT.FIELD"),
            GETTEXT_PO_V1: (ValidationOutcome.FAILED, 0, "PARSER.LIMIT.FIELD"),
            GETTEXT_POT_V1: (ValidationOutcome.FAILED, 0, "PARSER.LIMIT.FIELD"),
            TMX_LEVEL1_V1: (
                ValidationOutcome.SUCCESS,
                0,
                "PARSER.TMX.SEGMENT_LIMIT",
            ),
            NORMALIZED_TM_JSON_V1: (
                ValidationOutcome.FAILED,
                0,
                "PARSER.LIMIT.FIELD",
            ),
            TERMBASE_CSV_V1: (ValidationOutcome.FAILED, 0, "PARSER.LIMIT.FIELD"),
            TERMBASE_XLSX_V1: (ValidationOutcome.FAILED, 0, "PARSER.LIMIT.FIELD"),
        }
        record_expectations = {
            LOCALCAT_JSON_V1: 0,
            LINE_TEXT_V1: 1,
            GETTEXT_PO_V1: 1,
            GETTEXT_POT_V1: 1,
            TMX_LEVEL1_V1: 1,
            NORMALIZED_TM_JSON_V1: 0,
            TERMBASE_CSV_V1: 1,
            TERMBASE_XLSX_V1: 1,
        }
        materialization_expectations = {
            format_id: "PARSER.LIMIT.MATERIALIZATION"
            for format_id in expected_formats
        }
        self.assertEqual(set(field_expectations), expected_formats)
        self.assertEqual(set(record_expectations), expected_formats)
        self.assertEqual(set(materialization_expectations), expected_formats)

        for case in two_record_cases:
            original = descriptors[case.format_id]
            with self.subTest(format_id=case.format_id.value, axis="input"):
                input_profile = replace(
                    original.limit_profile,
                    max_input_bytes=len(case.payload) - 1,
                )
                input_descriptor = replace(original, limit_profile=input_profile)
                with self.assertRaises(ParserSourceError) as input_failure:
                    self._direct_snapshot(case, input_descriptor)
                self.assertEqual(
                    input_failure.exception.code,
                    input_expectations[case.format_id],
                )

            with self.subTest(format_id=case.format_id.value, axis="field"):
                field_profile = replace(
                    original.limit_profile,
                    max_decoded_field_chars=1,
                )
                field_descriptor = replace(original, limit_profile=field_profile)
                field_snapshot = self._direct_snapshot(case, field_descriptor)
                try:
                    with ExitStack() as stack:
                        field_report = validate(
                            self._direct_reader(case, field_descriptor, stack),
                            field_snapshot,
                            case.request(),
                        )
                finally:
                    field_snapshot.close()
                expected_outcome, expected_count, expected_code = field_expectations[
                    case.format_id
                ]
                self.assertIs(field_report.outcome, expected_outcome)
                self.assertEqual(field_report.provisional_record_count, expected_count)
                self.assertEqual(field_report.issues[-1].code, expected_code)
                if expected_outcome is ValidationOutcome.SUCCESS:
                    self.assertIsNotNone(field_report.terminal)
                else:
                    self.assertIsNone(field_report.terminal)

            with self.subTest(format_id=case.format_id.value, axis="record"):
                record_profile = replace(
                    original.limit_profile,
                    max_records=1,
                    max_materialized_records=1,
                )
                record_descriptor = replace(original, limit_profile=record_profile)
                record_snapshot = self._direct_snapshot(case, record_descriptor)
                try:
                    with ExitStack() as stack:
                        record_report = validate(
                            self._direct_reader(case, record_descriptor, stack),
                            record_snapshot,
                            case.request(),
                        )
                finally:
                    record_snapshot.close()
                self.assertIs(record_report.outcome, ValidationOutcome.FAILED)
                self.assertIsNone(record_report.terminal)
                self.assertEqual(
                    record_report.provisional_record_count,
                    record_expectations[case.format_id],
                )
                self.assertEqual(
                    record_report.issues[-1].code,
                    "PARSER.LIMIT.RECORD",
                )

            with self.subTest(
                format_id=case.format_id.value,
                axis="materialization",
            ):
                materialized_profile = replace(
                    original.limit_profile,
                    max_records=2,
                    max_materialized_records=1,
                )
                materialized_descriptor = replace(
                    original,
                    limit_profile=materialized_profile,
                )
                materialized_snapshot = self._direct_snapshot(
                    case,
                    materialized_descriptor,
                )
                try:
                    with ExitStack() as stack:
                        with self.assertRaises(ParserSessionError) as materialized_failure:
                            materialize(
                                self._direct_reader(
                                    case,
                                    materialized_descriptor,
                                    stack,
                                ),
                                materialized_snapshot,
                                case.request(),
                            )
                finally:
                    materialized_snapshot.close()
                self.assertEqual(
                    materialized_failure.exception.code,
                    materialization_expectations[case.format_id],
                )

    def test_gettext_real_metadata_is_bounded_for_po_and_pot(self) -> None:
        cases = (
            _InputCase(
                "metadata.po",
                GETTEXT_PO_V1,
                EffectivePurpose.PROJECT_DOCUMENT,
                b"# translator metadata longer than limit\n"
                b'msgid "A"\nmsgstr "B"\n',
            ),
            _InputCase(
                "metadata.pot",
                GETTEXT_POT_V1,
                EffectivePurpose.PROJECT_DOCUMENT,
                b"# translator metadata longer than limit\n"
                b'msgid "A"\nmsgstr ""\n',
            ),
        )
        originals = {
            GETTEXT_PO_V1: GETTEXT_PO_DESCRIPTOR,
            GETTEXT_POT_V1: GETTEXT_POT_DESCRIPTOR,
        }
        for case in cases:
            with self.subTest(format_id=case.format_id.value):
                original = originals[case.format_id]
                profile = replace(
                    original.limit_profile,
                    max_metadata_decoded_chars_per_container=8,
                    max_metadata_decoded_chars_total=8,
                )
                descriptor = replace(original, limit_profile=profile)
                snapshot = self._direct_snapshot(case, descriptor)
                try:
                    with ExitStack() as stack:
                        report = validate(
                            self._direct_reader(case, descriptor, stack),
                            snapshot,
                            case.request(),
                        )
                finally:
                    snapshot.close()
                self.assertIs(report.outcome, ValidationOutcome.FAILED)
                self.assertIsNone(report.terminal)
                self.assertEqual(
                    report.issues[-1].code,
                    "PARSER.LIMIT.METADATA",
                )

    def test_all_descriptors_publish_the_applicable_limit_and_issue_axes(
        self,
    ) -> None:
        cases = _valid_cases()
        self._assert_complete_builtin_matrix(cases)
        self.assertEqual(
            set(_EXPECTED_LIMIT_PROFILE_FACTS),
            {format_id for _purpose, format_id in _EXPECTED_BUILTIN_BINDINGS},
        )
        observed = {}
        for case in cases:
            with self.subTest(format_id=case.format_id.value):
                with self._open(case) as opened:
                    profile = opened.descriptor.limit_profile
                fact = _limit_profile_fact(profile)
                observed[case.format_id] = fact
                self.assertEqual(
                    fact,
                    _EXPECTED_LIMIT_PROFILE_FACTS[case.format_id],
                )
                self.assertEqual(
                    profile.declared_issue_codes,
                    tuple(sorted(set(profile.declared_issue_codes))),
                )
        self.assertEqual(observed, _EXPECTED_LIMIT_PROFILE_FACTS)

    def test_invalid_encoding_is_fatal_for_every_builtin_application_entry(
        self,
    ) -> None:
        invalid_cases = _invalid_encoding_cases()
        self._assert_complete_builtin_matrix(invalid_cases)
        for case in invalid_cases:
            with self.subTest(format_id=case.format_id.value):
                with self._open(case) as opened:
                    report = opened.validate()

                self.assertIs(report.outcome, ValidationOutcome.FAILED)
                self.assertIsNone(report.terminal)
                self.assertEqual(report.provisional_record_count, 0)
                self.assertEqual(
                    report.issues[-1].code,
                    "PARSER.SOURCE.ENCODING_FAILED",
                )

    def test_xlsx_inert_features_are_read_data_only_without_execution(self) -> None:
        payload = _xlsx_with_inert_features()
        self._assert_inert_xlsx_package(payload)
        case = _InputCase(
            "inert-features.xlsx",
            TERMBASE_XLSX_V1,
            EffectivePurpose.TERMBASE,
            payload,
        )
        import openpyxl

        actual_load_workbook = openpyxl.load_workbook
        observed_calls = []

        def observed_load_workbook(source, **kwargs):
            observed_calls.append(dict(kwargs))
            return actual_load_workbook(source, **kwargs)

        with mock.patch.object(
            socket,
            "socket",
            side_effect=AssertionError("network access attempted"),
        ), mock.patch(
            "parser_termbase_codec.importlib.import_module",
            wraps=__import__("importlib").import_module,
        ), mock.patch.object(
            openpyxl,
            "load_workbook",
            side_effect=observed_load_workbook,
        ), mock.patch(
            "subprocess.Popen",
            side_effect=AssertionError("process execution attempted"),
        ), mock.patch(
            "os.system",
            side_effect=AssertionError("shell execution attempted"),
        ):
            with self._open(case) as opened:
                result = opened.materialize()

        self.assertEqual(
            [(record.source, record.target) for record in result.records],
            [("A", "B")],
        )
        self.assertEqual(
            observed_calls,
            [
                {
                    "read_only": True,
                    "data_only": True,
                    "keep_links": False,
                    "keep_vba": False,
                }
            ],
        )
        serialized = repr(result.records)
        self.assertNotIn("WEBSERVICE", serialized)
        self.assertNotIn("invalid.example", serialized)

    def test_real_application_commit_faults_preserve_every_existing_target(
        self,
    ) -> None:
        project_target = self.root / "project.json"
        project_target.write_bytes(b"old-project")
        project = EditorProject(
            name="Project",
            segments=(EditorSegment(id="1", source="A", target="B"),),
        )
        before_names = {path.name for path in self.root.iterdir()}
        with mock.patch(
            "parser_source.os.replace",
            side_effect=OSError("injected replace failure"),
        ) as project_replace:
            with self.assertRaises(ProjectError):
                save_project(project, project_target)
        project_replace.assert_called_once()
        project_args, project_kwargs = project_replace.call_args
        self.assertEqual(len(project_args), 2)
        self.assertTrue(str(project_args[0]).startswith(".parser-"))
        self.assertTrue(str(project_args[0]).endswith(".tmp"))
        self.assertEqual(project_args[1], project_target.name)
        self.assertEqual(
            set(project_kwargs),
            {"src_dir_fd", "dst_dir_fd"},
        )
        self.assertEqual(
            project_kwargs["src_dir_fd"],
            project_kwargs["dst_dir_fd"],
        )
        self.assertEqual(project_target.read_bytes(), b"old-project")
        self.assertEqual({path.name for path in self.root.iterdir()}, before_names)

        term_source = self.root / "incoming.csv"
        term_source.write_bytes(b"Source,Target\nA,B\n")
        term_target = self.root / "managed.csv"
        term_target.write_text("Source,Target\nold,legacy\n", encoding="utf-8-sig")
        term_original = term_target.read_bytes()
        before_names = {path.name for path in self.root.iterdir()}
        with _record_real_resource_parser_bridge() as term_bridge, mock.patch(
            "resource_importer.os.replace",
            side_effect=OSError("injected replace failure"),
        ) as term_replace:
            term_report = import_termbase(term_source, term_target)
        self._assert_resource_parser_bridge(
            term_bridge,
            purpose=EffectivePurpose.TERMBASE,
            format_id=TERMBASE_CSV_V1,
            expected_count=2,
        )
        term_replace.assert_called_once()
        term_args, term_kwargs = term_replace.call_args
        self.assertEqual(term_kwargs, {})
        self.assertEqual(len(term_args), 2)
        self.assertIsInstance(term_args[0], Path)
        self.assertEqual(term_args[0].parent, term_target.resolve().parent)
        self.assertTrue(term_args[0].name.startswith(f".{term_target.name}."))
        self.assertTrue(term_args[0].name.endswith(".tmp"))
        self.assertEqual(term_args[1], term_target.resolve())
        self.assertFalse(term_report.succeeded)
        self.assertEqual(term_target.read_bytes(), term_original)
        self.assertEqual({path.name for path in self.root.iterdir()}, before_names)

        legacy_source = self.root / "incoming.tmx"
        legacy_source.write_bytes(
            b'<tmx><body><tu><tuv xml:lang="en-US"><seg>A</seg></tuv>'
            b'<tuv xml:lang="zh-CN"><seg>B</seg></tuv></tu></body></tmx>'
        )
        legacy_target = self.root / "legacy.jsonl"
        legacy_target.write_bytes(b'{"source":"old","target":"legacy"}\n')
        legacy_original = legacy_target.read_bytes()
        before_names = {path.name for path in self.root.iterdir()}
        with _record_real_resource_parser_bridge() as legacy_bridge, mock.patch(
            "resource_importer.os.replace",
            side_effect=OSError("injected replace failure"),
        ) as legacy_replace:
            legacy_report = import_tmx(
                legacy_source,
                legacy_target,
                "en-US",
                "zh-CN",
            )
        self._assert_resource_parser_bridge(
            legacy_bridge,
            purpose=EffectivePurpose.TRANSLATION_MEMORY,
            format_id=TMX_LEVEL1_V1,
        )
        legacy_replace.assert_called_once()
        legacy_args, legacy_kwargs = legacy_replace.call_args
        self.assertEqual(legacy_kwargs, {})
        self.assertEqual(len(legacy_args), 2)
        self.assertIsInstance(legacy_args[0], Path)
        self.assertEqual(legacy_args[0].parent, legacy_target.resolve().parent)
        self.assertTrue(legacy_args[0].name.startswith(f".{legacy_target.name}."))
        self.assertTrue(legacy_args[0].name.endswith(".tmp"))
        self.assertEqual(legacy_args[1], legacy_target.resolve())
        self.assertFalse(legacy_report.succeeded)
        self.assertEqual(legacy_target.read_bytes(), legacy_original)
        self.assertEqual({path.name for path in self.root.iterdir()}, before_names)

        canonical_root = self.root / "canonical"
        canonical_root.mkdir()
        canonical_target = _activate_resource(canonical_root)
        canonical_original = canonical_target.read_bytes()
        canonical_store = _store_for(canonical_target)
        original_count = canonical_store.canonical_revision().record_count
        canonical_source = canonical_root / "incoming.tmx"
        canonical_source.write_bytes(legacy_source.read_bytes())
        failure = SQLiteStoreLifecycleError(
            "STORE.TEST_BATCH_FAILURE",
            resource_id="tm.primary",
            generation=0,
            retryable=False,
        )
        with _record_real_resource_parser_bridge() as canonical_bridge, mock.patch.object(
            SQLiteTMStore,
            "append_batch",
            side_effect=failure,
        ) as append_batch:
            canonical_report = import_tmx(
                canonical_source,
                canonical_target,
                "en-US",
                "zh-CN",
            )
        self._assert_resource_parser_bridge(
            canonical_bridge,
            purpose=EffectivePurpose.TRANSLATION_MEMORY,
            format_id=TMX_LEVEL1_V1,
        )
        append_batch.assert_called_once()
        append_args, append_kwargs = append_batch.call_args
        self.assertEqual(append_args, ())
        self.assertEqual(
            set(append_kwargs),
            {
                "batch_id",
                "kind",
                "drafts",
                "source_digest",
                "source_path",
                "invalid_count",
                "duplicate_source_count",
            },
        )
        self.assertTrue(append_kwargs["batch_id"].startswith("import."))
        self.assertEqual(append_kwargs["kind"], "import")
        self.assertEqual(len(append_kwargs["drafts"]), 1)
        self.assertEqual(
            (
                append_kwargs["drafts"][0].source_raw,
                append_kwargs["drafts"][0].target_raw,
            ),
            ("A", "B"),
        )
        self.assertEqual(
            append_kwargs["source_digest"],
            hashlib.sha256(canonical_source.read_bytes()).hexdigest(),
        )
        self.assertEqual(append_kwargs["source_path"], canonical_source.resolve())
        self.assertEqual(append_kwargs["invalid_count"], 0)
        self.assertEqual(append_kwargs["duplicate_source_count"], 0)
        self.assertFalse(canonical_report.succeeded)
        self.assertEqual(canonical_target.read_bytes(), canonical_original)
        self.assertEqual(
            canonical_store.canonical_revision().record_count,
            original_count,
        )

    def test_dangerous_json_xml_and_xlsx_fail_through_the_application_surface(
        self,
    ) -> None:
        xlsx = io.BytesIO()
        with zipfile.ZipFile(xlsx, "w", compression=zipfile.ZIP_STORED) as archive:
            archive.writestr(
                "xl/workbook.xml",
                (
                    b'<!DOCTYPE workbook SYSTEM "https://invalid.example/secret.dtd">'
                    b"<workbook/>"
                ),
            )
        cases = (
            (
                _InputCase(
                    "unsafe.json",
                    LOCALCAT_JSON_V1,
                    EffectivePurpose.PROJECT_DOCUMENT,
                    b'[{"source":"LEAK-ME"}] trailing',
                ),
                "PARSER.SYNTAX.MALFORMED",
            ),
            (
                _InputCase(
                    "unsafe-memory.json",
                    NORMALIZED_TM_JSON_V1,
                    EffectivePurpose.TRANSLATION_MEMORY,
                    (b"[" * 65) + (b"]" * 65),
                ),
                "PARSER.LIMIT.DEPTH",
            ),
            (
                _InputCase(
                    "unsafe.tmx",
                    TMX_LEVEL1_V1,
                    EffectivePurpose.TRANSLATION_MEMORY,
                    (
                        b'<!DOCTYPE tmx SYSTEM "https://invalid.example/secret.dtd">'
                        b"<tmx><body/></tmx>"
                    ),
                ),
                "PARSER.TMX.UNSAFE_XML",
            ),
            (
                _InputCase(
                    "unsafe.xlsx",
                    TERMBASE_XLSX_V1,
                    EffectivePurpose.TERMBASE,
                    xlsx.getvalue(),
                ),
                "PARSER.TERMBASE.XML_DECLARATION_FORBIDDEN",
            ),
        )

        with mock.patch.object(
            socket,
            "socket",
            side_effect=AssertionError("network access attempted"),
        ):
            for case, expected_code in cases:
                with self.subTest(format_id=case.format_id.value):
                    with self._open(case) as opened:
                        report = opened.validate()

                    self.assertIs(report.outcome, ValidationOutcome.FAILED)
                    self.assertIsNone(report.terminal)
                    self.assertEqual(report.provisional_record_count, 0)
                    self.assertEqual(report.issues[-1].code, expected_code)
                    self.assertNotIn(
                        "LEAK-ME",
                        " ".join(issue.safe_summary for issue in report.issues),
                    )


if __name__ == "__main__":
    unittest.main()
