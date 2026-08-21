"""Contract tests for XLSX archive and OPC XML preflight."""

from __future__ import annotations

import builtins
import io
import struct
import warnings
import unittest
from unittest import mock
import zipfile


from parser_xlsx_support import (
    XlsxPreflightError,
    XlsxPreflightLimits,
    preflight_xlsx,
)


def _limits(**changes: int | float) -> XlsxPreflightLimits:
    values: dict[str, int | float] = {
        "max_archive_members": 4_096,
        "max_expanded_bytes": 256 * 1024 * 1024,
        "max_compression_ratio": 100.0,
        "max_xml_depth": 64,
    }
    values.update(changes)
    return XlsxPreflightLimits(**values)  # type: ignore[arg-type]


TEST_LIMITS = _limits()


def _preflight(
    source: bytes | bytearray | memoryview | io.BytesIO,
    limits: XlsxPreflightLimits = TEST_LIMITS,
):
    return preflight_xlsx(source, limits)


def _archive_bytes(
    members: list[tuple[str, bytes, int]] | None = None,
) -> bytes:
    if members is None:
        members = [
            (
                "[Content_Types].xml",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                    b'<Default Extension="rels" '
                    b'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                    b'<Default Extension="xml" ContentType="application/xml"/>'
                    b'<Override PartName="/xl/workbook.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
                    b'<Override PartName="/xl/worksheets/sheet1.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                    b"</Types>"
                ),
                zipfile.ZIP_STORED,
            ),
            (
                "_rels/.rels",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    b'<Relationship Id="rId1" '
                    b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
                    b'Target="xl/workbook.xml"/>'
                    b"</Relationships>"
                ),
                zipfile.ZIP_STORED,
            ),
            (
                "xl/workbook.xml",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                    b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                    b'<sheets><sheet name="Terms" sheetId="1" r:id="rId1"/></sheets>'
                    b"</workbook>"
                ),
                zipfile.ZIP_DEFLATED,
            ),
            (
                "xl/_rels/workbook.xml.rels",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    b'<Relationship Id="rId1" '
                    b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                    b'Target="worksheets/sheet1.xml"/>'
                    b"</Relationships>"
                ),
                zipfile.ZIP_STORED,
            ),
            (
                "xl/worksheets/sheet1.xml",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                    b'<sheetData><row r="1"><c r="A1" t="inlineStr"><is><t>Source</t></is></c>'
                    b'<c r="B1" t="inlineStr"><is><t>Target</t></is></c></row>'
                    b'<row r="2"><c r="A2" t="inlineStr"><is><t>Alpha</t></is></c>'
                    b'<c r="B2" t="inlineStr"><is><t>\xe7\x94\xb2</t></is></c></row></sheetData>'
                    b"</worksheet>"
                ),
                zipfile.ZIP_DEFLATED,
            ),
        ]
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        for name, payload, compression in members:
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = compression
            archive.writestr(info, payload)
    return output.getvalue()


def _data_descriptor_archive_bytes() -> bytes:
    class StreamingBuffer(io.BytesIO):
        def seekable(self) -> bool:
            return False

        def seek(self, _offset: int, _whence: int = 0) -> int:
            raise OSError("streaming ZIP output is not seekable")

    output = StreamingBuffer()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("streamed.xml", b"<streamed/>")
    return output.getvalue()


def _assert_code(test: unittest.TestCase, expected: str, call) -> XlsxPreflightError:
    with test.assertRaises(XlsxPreflightError) as raised:
        call()
    test.assertEqual(raised.exception.code, expected)
    return raised.exception


class XlsxArchivePreflightTests(unittest.TestCase):
    def test_module_is_stdlib_only_and_safe_workbook_returns_immutable_report(self) -> None:
        payload = _archive_bytes()
        source = io.BytesIO(payload)
        original_import = builtins.__import__

        def guarded_import(name, *args, **kwargs):
            if name == "openpyxl" or name.startswith("openpyxl."):
                raise AssertionError("XLSX preflight must run before openpyxl")
            return original_import(name, *args, **kwargs)

        with mock.patch("builtins.__import__", side_effect=guarded_import):
            report = _preflight(source)

        self.assertEqual(report.member_count, 5)
        self.assertEqual(report.crc_verified_member_count, report.member_count)
        self.assertEqual(report.xml_member_count, 5)
        self.assertEqual(
            report.xml_member_names,
            (
                "[Content_Types].xml",
                "_rels/.rels",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
                "xl/worksheets/sheet1.xml",
            ),
        )
        self.assertEqual(report.total_expanded_bytes, sum(item.file_size for item in report.members))
        self.assertEqual(source.tell(), 0)
        with self.assertRaises((AttributeError, TypeError)):
            report.member_count = 0  # type: ignore[misc]

    def test_safe_fixture_remains_readable_by_conditional_workbook_reader(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is an optional dependency")

        source = io.BytesIO(_archive_bytes())
        _preflight(source)
        workbook = load_workbook(
            source,
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
        try:
            self.assertEqual(workbook.active.title, "Terms")
            self.assertEqual(
                list(workbook.active.iter_rows(values_only=True)),
                [("Source", "Target"), ("Alpha", "甲")],
            )
        finally:
            workbook.close()

    def test_accepts_sealed_bytes_like_inputs(self) -> None:
        payload = _archive_bytes()
        expected = _preflight(payload)
        self.assertEqual(_preflight(bytearray(payload)), expected)
        self.assertEqual(_preflight(memoryview(payload)), expected)

    def test_limits_projection_is_required_from_the_caller(self) -> None:
        with self.assertRaises(TypeError):
            preflight_xlsx(_archive_bytes())  # type: ignore[call-arg]
        with self.assertRaises(TypeError):
            XlsxPreflightLimits()  # type: ignore[call-arg]

    def test_member_count_expansion_and_per_member_ratio_limits_are_injectable(self) -> None:
        payload = _archive_bytes(
            [
                ("one.xml", b"<one/>", zipfile.ZIP_STORED),
                ("two.xml", b"<two/>" + b" " * 256, zipfile.ZIP_DEFLATED),
            ]
        )
        _assert_code(
            self,
            "PARSER.LIMIT.ARCHIVE_MEMBER",
            lambda: _preflight(payload, _limits(max_archive_members=1)),
        )
        _assert_code(
            self,
            "PARSER.LIMIT.EXPANSION",
            lambda: _preflight(payload, _limits(max_expanded_bytes=16)),
        )
        error = _assert_code(
            self,
            "PARSER.LIMIT.COMPRESSION_RATIO",
            lambda: _preflight(payload, _limits(max_compression_ratio=2.0)),
        )
        self.assertEqual(error.member_name, "two.xml")
        self.assertGreater(error.observed, error.limit)

    def test_bad_zip_and_non_seekable_stream_fail_closed(self) -> None:
        _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_INVALID",
            lambda: _preflight(b"not a zip"),
        )
        _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_INVALID",
            lambda: _preflight(_archive_bytes()[:-12]),
        )

        class NonSeekable(io.BytesIO):
            def seekable(self) -> bool:
                return False

        _assert_code(
            self,
            "PARSER.XLSX.SOURCE_NOT_SEEKABLE",
            lambda: _preflight(NonSeekable(_archive_bytes())),
        )

    def test_cursor_restore_failure_is_a_structured_failure(self) -> None:
        payload = _archive_bytes()
        sentinel = len(payload) + 17

        class RestoreFailing(io.BytesIO):
            armed = False

            def seek(self, offset: int, whence: int = 0) -> int:
                if self.armed and whence == 0 and offset == sentinel:
                    raise OSError("native restore detail must not leak")
                return super().seek(offset, whence)

        source = RestoreFailing(payload)
        io.BytesIO.seek(source, sentinel)
        source.armed = True
        error = _assert_code(
            self,
            "PARSER.XLSX.SOURCE_RESTORE_FAILED",
            lambda: _preflight(source),
        )
        self.assertNotIn("native restore", str(error))

    def test_cursor_restore_must_reach_the_original_offset(self) -> None:
        payload = _archive_bytes()
        sentinel = len(payload) + 17

        class LyingRestore(io.BytesIO):
            armed = False

            def seek(self, offset: int, whence: int = 0) -> int:
                if self.armed and whence == 0 and offset == sentinel:
                    return super().seek(0)
                return super().seek(offset, whence)

        source = LyingRestore(payload)
        io.BytesIO.seek(source, sentinel)
        source.armed = True
        _assert_code(
            self,
            "PARSER.XLSX.SOURCE_RESTORE_FAILED",
            lambda: _preflight(source),
        )

    def test_failed_offset_zero_probe_restores_entry_offset(self) -> None:
        class PartiallyMovingCursor(io.BytesIO):
            first_zero_seek = True

            def seek(self, offset: int, whence: int = 0) -> int:
                position = super().seek(offset, whence)
                if self.first_zero_seek and whence == 0 and offset == 0:
                    self.first_zero_seek = False
                    return 1
                return position

        source = PartiallyMovingCursor(_archive_bytes())
        source.seek(7)
        error = _assert_code(
            self,
            "PARSER.XLSX.SOURCE_NOT_SEEKABLE",
            lambda: _preflight(source),
        )
        self.assertEqual(source.tell(), 7)
        self.assertEqual(str(error), "PARSER.XLSX.SOURCE_NOT_SEEKABLE")

    def test_bad_local_header_in_non_xml_member_still_fails_closed(self) -> None:
        payload = bytearray(
            _archive_bytes(
                [
                    ("safe.xml", b"<safe/>", zipfile.ZIP_STORED),
                    ("xl/media/blob.bin", b"binary", zipfile.ZIP_STORED),
                ]
            )
        )
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            offset = archive.getinfo("xl/media/blob.bin").header_offset
        self.assertEqual(struct.unpack_from("<I", payload, offset)[0], 0x04034B50)
        struct.pack_into("<I", payload, offset, 0xDEADBEEF)

        error = _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_INVALID",
            lambda: _preflight(payload),
        )
        self.assertEqual(error.member_name, "xl/media/blob.bin")

    def test_every_binary_member_is_fully_decompressed_and_crc_checked(self) -> None:
        payload = bytearray(
            _archive_bytes(
                [
                    ("safe.xml", b"<safe/>", zipfile.ZIP_STORED),
                    ("xl/media/blob.bin", b"binary-payload", zipfile.ZIP_STORED),
                ]
            )
        )
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            info = archive.getinfo("xl/media/blob.bin")
            local = payload[info.header_offset : info.header_offset + 30]
            name_length, extra_length = struct.unpack_from("<HH", local, 26)
            data_start = info.header_offset + 30 + name_length + extra_length
        payload[data_start] ^= 0x01

        error = _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_INVALID",
            lambda: _preflight(payload),
        )
        self.assertEqual(error.member_name, "xl/media/blob.bin")

    def test_local_header_crc_and_sizes_must_match_central_directory(self) -> None:
        for field_offset in (14, 18, 22):
            with self.subTest(local_header_field=field_offset):
                payload = bytearray(_archive_bytes())
                with zipfile.ZipFile(io.BytesIO(payload)) as archive:
                    info = archive.getinfo("xl/workbook.xml")
                original = struct.unpack_from("<I", payload, info.header_offset + field_offset)[0]
                struct.pack_into(
                    "<I",
                    payload,
                    info.header_offset + field_offset,
                    original ^ 0x01,
                )
                error = _assert_code(
                    self,
                    "PARSER.XLSX.ARCHIVE_INVALID",
                    lambda payload=payload: _preflight(payload),
                )
                self.assertEqual(error.member_name, "xl/workbook.xml")

    def test_data_descriptor_members_are_explicitly_rejected_before_decompression(self) -> None:
        valid = _data_descriptor_archive_bytes()
        with zipfile.ZipFile(io.BytesIO(valid)) as archive:
            info = archive.getinfo("streamed.xml")
            self.assertTrue(info.flag_bits & 0x08)
            local = valid[info.header_offset : info.header_offset + 30]
            name_length, extra_length = struct.unpack_from("<HH", local, 26)
            descriptor_offset = (
                info.header_offset + 30 + name_length + extra_length + info.compress_size
            )
        self.assertEqual(valid[descriptor_offset : descriptor_offset + 4], b"PK\x07\x08")
        zeroed = bytearray(valid)
        zeroed[descriptor_offset : descriptor_offset + 16] = b"\x00" * 16

        for payload in (valid, zeroed):
            with self.subTest(descriptor_zeroed=payload is zeroed):
                error = _assert_code(
                    self,
                    "PARSER.XLSX.ARCHIVE_DATA_DESCRIPTOR_UNSUPPORTED",
                    lambda payload=payload: _preflight(payload),
                )
                self.assertEqual(error.member_name, "streamed.xml")

    def test_corrupt_lzma_member_maps_to_stable_body_safe_archive_code(self) -> None:
        payload = bytearray(
            _archive_bytes(
                [("xl/media/blob.bin", b"A" * 2_048, zipfile.ZIP_LZMA)]
            )
        )
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            info = archive.getinfo("xl/media/blob.bin")
            local = payload[info.header_offset : info.header_offset + 30]
            name_length, extra_length = struct.unpack_from("<HH", local, 26)
            data_start = info.header_offset + 30 + name_length + extra_length
        payload[data_start + info.compress_size // 2] ^= 0xFF

        error = _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_INVALID",
            lambda: _preflight(payload),
        )
        self.assertEqual(error.member_name, "xl/media/blob.bin")
        self.assertEqual(str(error), "PARSER.XLSX.ARCHIVE_INVALID; member=xl/media/blob.bin")

    def test_duplicate_and_unsafe_member_names_are_rejected(self) -> None:
        output = io.BytesIO()
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            with zipfile.ZipFile(output, "w") as archive:
                archive.writestr("same.xml", "<first/>")
                archive.writestr("same.xml", "<second/>")
        duplicate = _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_MEMBER_DUPLICATE",
            lambda: _preflight(output.getvalue()),
        )
        self.assertEqual(duplicate.member_name, "same.xml")

        for unsafe in ("/absolute.xml", "../escape.xml", "a/../escape.xml", "a\\escape.xml", "a//b.xml"):
            with self.subTest(name=unsafe):
                payload = _archive_bytes([(unsafe, b"<root/>", zipfile.ZIP_STORED)])
                error = _assert_code(
                    self,
                    "PARSER.XLSX.ARCHIVE_MEMBER_NAME_UNSAFE",
                    lambda payload=payload: _preflight(payload),
                )
                self.assertEqual(error.member_name, unsafe)

    def test_member_name_in_failure_is_bounded_and_body_safe(self) -> None:
        raw_name = "n" * (65_535 - len(".xml")) + ".xml"
        payload = _archive_bytes([(raw_name, b"<broken>", zipfile.ZIP_STORED)])
        error = _assert_code(
            self,
            "PARSER.SYNTAX.MALFORMED",
            lambda: _preflight(payload),
        )
        self.assertLessEqual(len(error.member_name), 160)
        self.assertLessEqual(len(str(error)), 220)
        self.assertNotIn(raw_name, str(error))

        injected = "bad\nname.xml"
        injected_error = _assert_code(
            self,
            "PARSER.XLSX.ARCHIVE_MEMBER_NAME_UNSAFE",
            lambda: _preflight(
                _archive_bytes([(injected, b"<root/>", zipfile.ZIP_STORED)])
            ),
        )
        self.assertNotIn("\n", str(injected_error))
        self.assertEqual(injected_error.member_name, "bad?name.xml")


class XlsxOpcXmlPreflightTests(unittest.TestCase):
    def test_dtd_entity_and_external_entity_are_forbidden(self) -> None:
        cases = {
            "dtd.xml": b"<!DOCTYPE root><root/>",
            "entity.xml": b"<!DOCTYPE root [<!ENTITY local 'x'>]><root>&local;</root>",
            "external.xml": b"<!DOCTYPE root [<!ENTITY ext SYSTEM 'https://invalid.example/x'>]><root>&ext;</root>",
            "parameter.xml": (
                b"<!DOCTYPE root [<!ENTITY % ext SYSTEM "
                b"'https://invalid.example/p'>%ext;]><root/>"
            ),
        }
        for name, xml in cases.items():
            with self.subTest(name=name):
                error = _assert_code(
                    self,
                    "PARSER.TERMBASE.XML_DECLARATION_FORBIDDEN",
                    lambda name=name, xml=xml: _preflight(
                        _archive_bytes([(name, xml, zipfile.ZIP_STORED)])
                    ),
                )
                self.assertEqual(error.member_name, name)

    def test_invalid_xml_and_invalid_utf8_have_distinct_stable_codes(self) -> None:
        malformed = _assert_code(
            self,
            "PARSER.SYNTAX.MALFORMED",
            lambda: _preflight(
                _archive_bytes([("broken.xml", b"<root>", zipfile.ZIP_STORED)])
            ),
        )
        self.assertEqual(malformed.member_name, "broken.xml")

        encoding = _assert_code(
            self,
            "PARSER.SOURCE.ENCODING_FAILED",
            lambda: _preflight(
                _archive_bytes([("encoding.xml", b"<root>\xff</root>", zipfile.ZIP_STORED)])
            ),
        )
        self.assertEqual(encoding.member_name, "encoding.xml")

    def test_every_xml_and_relationship_member_is_checked(self) -> None:
        for second_name in ("customXml/item2.xml", "customXml/_rels/item1.xml.rels", "customXml/ITEM.XML"):
            with self.subTest(member=second_name):
                payload = _archive_bytes(
                    [
                        ("customXml/item1.xml", b"<first/>", zipfile.ZIP_STORED),
                        (second_name, b"<broken>", zipfile.ZIP_STORED),
                    ]
                )
                error = _assert_code(
                    self,
                    "PARSER.SYNTAX.MALFORMED",
                    lambda payload=payload: _preflight(payload),
                )
                self.assertEqual(error.member_name, second_name)

    def test_non_xml_binary_members_are_enumerated_but_not_decoded_or_parsed(self) -> None:
        payload = _archive_bytes(
            [
                ("xl/workbook.xml", b"<workbook/>", zipfile.ZIP_STORED),
                ("xl/vbaProject.bin", b"\xff\x00<!DOCTYPE not-xml>", zipfile.ZIP_STORED),
                ("xl/media/image.bin", b"\x00\xff\xfe", zipfile.ZIP_DEFLATED),
            ]
        )
        report = _preflight(payload)
        self.assertEqual(report.member_count, 3)
        self.assertEqual(report.xml_member_names, ("xl/workbook.xml",))

    def test_structure_depth_is_bounded(self) -> None:
        payload = _archive_bytes(
            [("deep.xml", b"<a><b><c/></b></a>", zipfile.ZIP_STORED)]
        )
        error = _assert_code(
            self,
            "PARSER.LIMIT.STRUCTURE_DEPTH",
            lambda: _preflight(payload, _limits(max_xml_depth=2)),
        )
        self.assertEqual(error.member_name, "deep.xml")


if __name__ == "__main__":
    unittest.main()
