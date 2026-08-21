from __future__ import annotations

import csv
from dataclasses import replace
import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from editor_contracts import (
    TermbaseImportHeaderMode,
    TermbaseImportSelection,
)
from parser_composition import create_parser_application_surface
from resource_importer import (
    ImportFailure,
    import_termbase,
    import_tmx,
    preview_termbase_import,
    read_legacy_termbase_import,
    upsert_term,
)


class ResourceImporterTest(unittest.TestCase):
    @staticmethod
    def _selection(
        preview,
        *,
        source: int,
        target: int,
        header_mode: TermbaseImportHeaderMode,
    ) -> TermbaseImportSelection:
        return TermbaseImportSelection(
            source_zero_based_index=source,
            target_zero_based_index=target,
            header_mode=header_mode,
            preview_column_count=len(preview.columns),
            preview_source_identity=preview.source_identity,
        )

    def _write_tmx(self, path: Path, body: str) -> None:
        path.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<tmx version="1.4"><header srclang="en-US"/><body>'
            f"{body}</body></tmx>",
            encoding="utf-8",
        )

    def _read_jsonl(self, path: Path) -> list[dict[str, object]]:
        return [
            json.loads(line)
            for line in path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]

    def _read_terms(self, path: Path) -> list[list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.reader(handle))

    def test_imports_matecat_tmx_with_locale_normalization_and_last_write_wins(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "matecat.tmx"
            target = root / "tm.jsonl"
            target.write_text(
                json.dumps({"source": "Existing", "target": "旧值"}, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )
            self._write_tmx(
                source,
                """
                <tu><tuv xml:lang="en_US"><seg>Existing</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>新值</seg></tuv></tu>
                <tu><tuv xml:lang="en-GB"><seg>New source</seg></tuv>
                    <tuv xml:lang="zh-Hans"><seg>新译文</seg></tuv></tu>
                <tu><tuv xml:lang="en-US"><seg>New source</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>最终译文</seg></tuv></tu>
                """,
            )

            report = import_tmx(source, target, "EN-us", "zh_cn")
            records = self._read_jsonl(target)

        self.assertEqual(report.imported, 2)
        self.assertEqual(report.overwritten, 2)
        self.assertEqual(report.skipped, 0)
        self.assertEqual(report.errors, ())
        self.assertEqual(
            {record["source"]: record["target"] for record in records},
            {"Existing": "新值", "New source": "最终译文"},
        )

    def test_tmx_skips_missing_pairs_and_inline_xml_but_imports_valid_units(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "partial.tmx"
            target = root / "tm.jsonl"
            target.write_text("", encoding="utf-8")
            self._write_tmx(
                source,
                """
                <tu><tuv xml:lang="en-US"><seg>Valid</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>有效</seg></tuv></tu>
                <tu><tuv xml:lang="en-US"><seg>Missing target</seg></tuv></tu>
                <tu><tuv xml:lang="en-US"><seg>Has <ph id="1"/> tag</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>有标签</seg></tuv></tu>
                """,
            )

            report = import_tmx(source, target, "en-US", "zh-CN")
            records = self._read_jsonl(target)

        self.assertEqual(report.imported, 1)
        self.assertEqual(report.skipped, 2)
        self.assertTrue(any("inline" in error.lower() for error in report.errors))
        self.assertTrue(
            any(
                error.startswith("PARSER.TMX.LOCALE_PAIR_MISSING:")
                for error in report.errors
            )
        )
        self.assertTrue(
            any(
                error.startswith("PARSER.TMX.INLINE_XML_UNSUPPORTED:")
                for error in report.errors
            )
        )
        self.assertEqual(records[0]["source"], "Valid")

    def test_tmx_skips_oversized_segments(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "long-segment.tmx"
            target = root / "tm.jsonl"
            target.write_text("", encoding="utf-8")
            self._write_tmx(
                source,
                """
                <tu><tuv xml:lang="en-US"><seg>Valid</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>有效</seg></tuv></tu>
                <tu><tuv xml:lang="en-US"><seg>Too long</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>超长</seg></tuv></tu>
                """,
            )

            from parser_tmx_codec import TMX_CODEC_DESCRIPTOR

            bounded_profile = replace(
                TMX_CODEC_DESCRIPTOR.limit_profile,
                max_decoded_field_chars=5,
            )
            bounded_descriptor = replace(
                TMX_CODEC_DESCRIPTOR,
                limit_profile=bounded_profile,
            )
            with (
                patch("parser_tmx_codec.TMX_CODEC_DESCRIPTOR", bounded_descriptor),
                patch("parser_composition._TMX_CODEC_DESCRIPTOR", bounded_descriptor),
            ):
                report = import_tmx(source, target, "en-US", "zh-CN")

        self.assertEqual(report.imported, 1)
        self.assertEqual(report.skipped, 1)
        self.assertTrue(
            any(
                error.startswith("PARSER.TMX.SEGMENT_LIMIT:")
                for error in report.errors
            )
        )

    def test_destructive_tmx_inputs_never_change_target(self) -> None:
        cases = {
            "dtd": (
                b'<?xml version="1.0"?><!DOCTYPE tmx SYSTEM "tmx14.dtd">'
                b'<tmx><body/></tmx>'
            ),
            "entity": b'<!ENTITY unsafe "value"><tmx><body/></tmx>',
            "malformed": b"<tmx><body>",
            "wrong-language": (
                b'<tmx><body><tu><tuv xml:lang="fr"><seg>Bonjour</seg></tuv>'
                b'<tuv xml:lang="de"><seg>Hallo</seg></tuv></tu></body></tmx>'
            ),
        }
        for label, payload in cases.items():
            with self.subTest(label=label), tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                source = root / f"{label}.tmx"
                target = root / "tm.jsonl"
                source.write_bytes(payload)
                original = b'{"source":"Keep","target":"\xe4\xbf\x9d\xe7\x95\x99"}\n'
                target.write_bytes(original)

                report = import_tmx(source, target, "en-US", "zh-CN")

                self.assertTrue(report.errors)
                self.assertEqual(target.read_bytes(), original)

    def test_ambiguous_base_locale_fallback_preserves_target(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "ambiguous.tmx"
            target = root / "tm.jsonl"
            target.write_bytes(b"keep")
            self._write_tmx(
                source,
                """
                <tu><tuv xml:lang="en-GB"><seg>Colour</seg></tuv>
                    <tuv xml:lang="en-AU"><seg>Colour AU</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>颜色</seg></tuv></tu>
                """,
            )

            report = import_tmx(source, target, "en", "zh-CN")
            target_bytes = target.read_bytes()

        self.assertTrue(report.errors)
        self.assertEqual(target_bytes, b"keep")

    def test_invalid_locale_selection_is_body_safe_and_non_committing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "valid.tmx"
            target = root / "tm.jsonl"
            target.write_bytes(b"keep")
            self._write_tmx(
                source,
                '<tu><tuv xml:lang="en-US"><seg>Private source</seg></tuv>'
                '<tuv xml:lang="zh-CN"><seg>私密译文</seg></tuv></tu>',
            )

            report = import_tmx(source, target, "en-US\nPrivate source", "zh-CN")

            self.assertEqual(target.read_bytes(), b"keep")

        self.assertEqual(report.imported, 0)
        self.assertEqual(report.skipped, 0)
        self.assertEqual(report.overwritten, 0)
        self.assertEqual(
            report.errors,
            (
                "PARSER.TMX.LOCALE_SELECTION_INVALID: "
                "TMX source and target locale selection is invalid",
            ),
        )
        self.assertNotIn("Private source", report.errors[0])
        self.assertNotIn("私密译文", report.errors[0])

    def test_valid_tmx_does_not_replace_invalid_existing_tm(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "valid.tmx"
            target = root / "tm.jsonl"
            original = b"{not-json\n"
            target.write_bytes(original)
            self._write_tmx(
                source,
                """
                <tu><tuv xml:lang="en-US"><seg>Valid</seg></tuv>
                    <tuv xml:lang="zh-CN"><seg>有效</seg></tuv></tu>
                """,
            )

            report = import_tmx(source, target, "en-US", "zh-CN")
            target_bytes = target.read_bytes()

        self.assertTrue(report.errors)
        self.assertEqual(target_bytes, original)

    def test_size_limit_rejects_before_replacing_target(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "large.tmx"
            target = root / "tm.jsonl"
            source.write_bytes(b"x" * 32)
            target.write_bytes(b"keep")

            from parser_tmx_codec import TMX_CODEC_DESCRIPTOR

            bounded_profile = replace(
                TMX_CODEC_DESCRIPTOR.limit_profile,
                max_input_bytes=16,
            )
            bounded_descriptor = replace(
                TMX_CODEC_DESCRIPTOR,
                limit_profile=bounded_profile,
            )
            with (
                patch("parser_tmx_codec.TMX_CODEC_DESCRIPTOR", bounded_descriptor),
                patch("parser_composition._TMX_CODEC_DESCRIPTOR", bounded_descriptor),
            ):
                report = import_tmx(source, target, "en-US", "zh-CN")

            self.assertTrue(any("PARSER.LIMIT.INPUT" in error for error in report.errors))
            self.assertEqual(target.read_bytes(), b"keep")

    def test_resource_facades_use_parser_surface_for_tmx_termbase_and_upsert(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tmx = root / "one.tmx"
            tm_target = root / "tm.jsonl"
            terms = root / "terms.csv"
            term_target = root / "managed.csv"
            self._write_tmx(
                tmx,
                '<tu><tuv xml:lang="en-US"><seg>Alpha</seg></tuv>'
                '<tuv xml:lang="zh-CN"><seg>甲</seg></tuv></tu>',
            )
            terms.write_text("Source,Target\nAlpha,甲\n", encoding="utf-8-sig")
            term_target.write_text("Keep,stable\n", encoding="utf-8-sig")

            with patch(
                "resource_importer.create_parser_application_surface",
                wraps=create_parser_application_surface,
            ) as surface_factory:
                tmx_report = import_tmx(tmx, tm_target, "en-US", "zh-CN")
                term_report = import_termbase(terms, term_target)
                upsert_report = upsert_term(term_target, "Fresh", "new")

        self.assertTrue(tmx_report.succeeded)
        self.assertTrue(term_report.succeeded)
        self.assertTrue(upsert_report.succeeded)
        self.assertGreaterEqual(surface_factory.call_count, 4)

    def test_resource_importer_has_no_private_tmx_or_termbase_parser(self) -> None:
        import resource_importer

        for name in (
            "_read_tmx_snapshot",
            "_parse_tmx",
            "_normalize_locale",
            "_select_locale",
            "_tuv_language",
            "_direct_child",
            "_local_name",
            "_read_termbase_rows",
            "_read_csv_rows",
            "_collect_terms",
            "_is_header",
        ):
            with self.subTest(name=name):
                self.assertFalse(hasattr(resource_importer, name))

    def test_termbase_preview_projects_bounded_codec_owned_csv_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.csv"
            source.write_text(
                "Target,Notes,Source\n甲,first,Alpha\n",
                encoding="utf-8-sig",
            )

            preview = preview_termbase_import(source)
            source_size = source.stat().st_size

        self.assertEqual(preview.format_name, "csv")
        self.assertEqual(
            tuple(
                (column.zero_based_index, column.header_candidate)
                for column in preview.columns
            ),
            ((0, "Target"), (1, "Notes"), (2, "Source")),
        )
        self.assertEqual(preview.total_column_count, 3)
        self.assertFalse(preview.columns_truncated)
        self.assertFalse(preview.legacy_header_detected)
        self.assertIsNone(preview.active_sheet_name)
        self.assertEqual(preview.columns[0].header_original_char_count, 6)
        self.assertFalse(preview.columns[0].header_truncated)
        self.assertEqual(preview.source_identity.original_size, source_size)
        self.assertEqual(
            preview.source_identity.original_size,
            preview.source_identity.byte_count,
        )

    def test_explicit_selection_maps_indices_and_header_mode_through_parser(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            headed = root / "headed.csv"
            headed.write_text(
                "Target,Notes,Source\n甲,first,Alpha\n乙,second,Beta\n",
                encoding="utf-8-sig",
            )
            headed_preview = preview_termbase_import(headed)
            headed_rows, headed_skipped = read_legacy_termbase_import(
                headed,
                self._selection(
                    headed_preview,
                    source=2,
                    target=0,
                    header_mode=TermbaseImportHeaderMode.FIRST_ROW,
                ),
            )

            headerless = root / "headerless.csv"
            headerless.write_text(
                "甲,first,Alpha\n乙,second,Beta\n",
                encoding="utf-8-sig",
            )
            headerless_preview = preview_termbase_import(headerless)
            headerless_rows, headerless_skipped = read_legacy_termbase_import(
                headerless,
                self._selection(
                    headerless_preview,
                    source=2,
                    target=0,
                    header_mode=TermbaseImportHeaderMode.NO_HEADER,
                ),
            )

        self.assertEqual(
            tuple((row.source, row.target, row.input_ordinal) for row in headed_rows),
            (("Alpha", "甲", 1), ("Beta", "乙", 2)),
        )
        self.assertEqual(headed_skipped, 1)
        self.assertEqual(
            tuple(
                (row.source, row.target, row.input_ordinal)
                for row in headerless_rows
            ),
            (("Alpha", "甲", 0), ("Beta", "乙", 1)),
        )
        self.assertEqual(headerless_skipped, 0)

    def test_xlsx_preview_projects_only_the_active_sheet_identity(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.xlsx"
            workbook = Workbook()
            inactive = workbook.active
            inactive.title = "Ignored"
            inactive.append(["Wrong", "Wrong target"])
            active = workbook.create_sheet("Selected")
            active.append(["Target", "Source"])
            active.append(["甲", "Alpha"])
            workbook.active = workbook.sheetnames.index("Selected")
            workbook.save(source)
            workbook.close()

            preview = preview_termbase_import(source)

        self.assertEqual(preview.format_name, "xlsx")
        self.assertEqual(preview.active_sheet_name, "Selected")
        self.assertEqual(
            tuple(column.header_candidate for column in preview.columns),
            ("Target", "Source"),
        )

    def test_stale_preview_fails_before_streaming_any_record(self) -> None:
        from parser_composition import OpenedParserInput

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.csv"
            source.write_text("Source,Target\nAlpha,甲\n", encoding="utf-8-sig")
            preview = preview_termbase_import(source)
            selection = self._selection(
                preview,
                source=0,
                target=1,
                header_mode=TermbaseImportHeaderMode.FIRST_ROW,
            )
            status = source.stat()
            os.utime(
                source,
                ns=(status.st_atime_ns, status.st_mtime_ns + 1_000_000_000),
            )

            with patch.object(
                OpenedParserInput,
                "stream",
                side_effect=AssertionError("stale import attempted to stream"),
            ):
                with self.assertRaisesRegex(ImportFailure, "PARSER\\.SOURCE\\.STALE"):
                    _ = read_legacy_termbase_import(source, selection)

    def test_current_preview_width_must_match_before_streaming(self) -> None:
        from parser_composition import OpenedParserInput

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.csv"
            source.write_text(
                "Source,Target,Notes\nAlpha,甲,first\n",
                encoding="utf-8-sig",
            )
            preview = preview_termbase_import(source)
            forged = TermbaseImportSelection(
                source_zero_based_index=0,
                target_zero_based_index=1,
                header_mode=TermbaseImportHeaderMode.FIRST_ROW,
                preview_column_count=2,
                preview_source_identity=preview.source_identity,
            )

            with patch.object(
                OpenedParserInput,
                "stream",
                side_effect=AssertionError("width mismatch attempted to stream"),
            ) as stream:
                with self.assertRaisesRegex(
                    ImportFailure,
                    "PARSER\\.TERMBASE\\.COLUMN_SELECTION_INVALID",
                ):
                    _ = read_legacy_termbase_import(source, forged)
            stream.assert_not_called()

    def test_preview_does_not_swallow_unexpected_programmer_fault(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.csv"
            source.write_text("Source,Target\nAlpha,甲\n", encoding="utf-8-sig")
            surface = create_parser_application_surface()
            with (
                patch(
                    "resource_importer.create_parser_application_surface",
                    return_value=surface,
                ),
                patch.object(
                    type(surface),
                    "preview_termbase_columns",
                    side_effect=AssertionError("injected preview programmer fault"),
                ),
            ):
                with self.assertRaisesRegex(
                    AssertionError,
                    "injected preview programmer fault",
                ):
                    _ = preview_termbase_import(source)

    def test_codec_preview_programmer_fault_remains_observable(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.csv"
            source.write_text("Source,Target\nAlpha,甲\n", encoding="utf-8-sig")
            with patch(
                "parser_termbase_codec.CsvTermbaseCodec.preview_columns",
                side_effect=AssertionError("codec programmer fault"),
            ):
                with self.assertRaisesRegex(AssertionError, "codec programmer fault"):
                    _ = preview_termbase_import(source)

    def test_xlsx_preview_distinguishes_input_failure_from_programmer_fault(self) -> None:
        from openpyxl.utils.exceptions import InvalidFileException

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "terms.xlsx"
            workbook = Workbook()
            workbook.active.append(["Source", "Target"])
            workbook.save(source)
            workbook.close()

            with patch(
                "openpyxl.load_workbook",
                side_effect=InvalidFileException("input detail"),
            ):
                with self.assertRaisesRegex(
                    ImportFailure,
                    "PARSER\\.SYNTAX\\.MALFORMED",
                ) as malformed:
                    _ = preview_termbase_import(source)
            self.assertNotIn("input detail", str(malformed.exception))

            with patch(
                "openpyxl.load_workbook",
                side_effect=AssertionError("xlsx programmer fault"),
            ):
                with self.assertRaisesRegex(AssertionError, "xlsx programmer fault"):
                    _ = preview_termbase_import(source)

    def test_imports_csv_headers_empty_rows_and_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "terms.csv"
            target = root / "terms-target.csv"
            source.write_text(
                "Source,Target,Notes\nEngine,引擎,one\n,\nEngine,发动机,two\nMemory,记忆库,\n",
                encoding="utf-8-sig",
            )
            target.write_text("Engine,旧译\nExisting,保留\n", encoding="utf-8-sig")

            report = import_termbase(source, target)
            rows = self._read_terms(target)

        self.assertEqual(report.imported, 2)
        self.assertEqual(report.skipped, 2)
        self.assertEqual(report.overwritten, 2)
        self.assertEqual(report.errors, ())
        self.assertEqual(dict(rows), {"Engine": "发动机", "Existing": "保留", "Memory": "记忆库"})
        self.assertNotIn(["Source", "Target"], rows)

    def test_imports_xlsx_first_two_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "terms.xlsx"
            target = root / "terms.csv"
            target.write_text("", encoding="utf-8-sig")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Source term", "Target term", "Ignored"])
            sheet.append(["office", "办公室", "note"])
            sheet.append(["memory", "记忆库", None])
            workbook.save(source)
            workbook.close()

            report = import_termbase(source, target)
            rows = self._read_terms(target)

        self.assertEqual(report.imported, 2)
        self.assertEqual(report.skipped, 1)
        self.assertEqual(report.errors, ())
        self.assertEqual(rows, [["office", "办公室"], ["memory", "记忆库"]])

    def test_invalid_termbase_preserves_original_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "terms.txt"
            target = root / "terms.csv"
            source.write_text("one,two", encoding="utf-8")
            target.write_bytes(b"\xef\xbb\xbfKeep,\xe4\xbf\x9d\xe7\x95\x99\r\n")
            original = target.read_bytes()

            report = import_termbase(source, target)

            self.assertTrue(report.errors)
            self.assertEqual(target.read_bytes(), original)

    def test_corrupt_xlsx_returns_error_and_preserves_original_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "broken.xlsx"
            target = root / "terms.csv"
            source.write_bytes(b"not a zip workbook")
            target.write_bytes(b"\xef\xbb\xbfKeep,\xe4\xbf\x9d\xe7\x95\x99\r\n")
            original = target.read_bytes()

            report = import_termbase(source, target)

            self.assertTrue(report.errors)
            self.assertEqual(target.read_bytes(), original)

    def test_upsert_term_is_atomic_and_rejects_invalid_existing_target(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            target = root / "terms.csv"
            target.write_text("office,旧译\n", encoding="utf-8-sig")

            report = upsert_term(target, "office", "办公室")
            rows = self._read_terms(target)
            target.write_bytes(b"broken-row")
            original = target.read_bytes()
            failed = upsert_term(target, "memory", "记忆库")

            self.assertEqual(report.imported, 1)
            self.assertEqual(report.overwritten, 1)
            self.assertEqual(rows, [["office", "办公室"]])
            self.assertTrue(failed.errors)
            self.assertEqual(target.read_bytes(), original)


if __name__ == "__main__":
    unittest.main()
